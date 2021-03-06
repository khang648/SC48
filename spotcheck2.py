#!/usr/bin/python3
########################################################### IMPORT MODULES - START #################################################################
from tkinter import *
from tkinter import messagebox
import time
from time import sleep, gmtime, strftime
from picamera import PiCamera
import cv2
import numpy as np
from tkinter import filedialog
from PIL import ImageTk, Image
import serial
from functools import partial
import math
from fractions import Fraction
from threading import *
import os
from tkinter import ttk
import awesometkinter as atk
import tkinter.font as font
import openpyxl
import subprocess
import shutil
import RPi.GPIO as GPIO
from ftplib import FTP
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Protection
from openpyxl.styles.borders import Border, Side
from openpyxl.drawing.image import Image as Img
from datetime import *
############################################################ IMPORT MODULES - END ##################################################################

########################################################## GLOBAL VARIABLE - START #################################################################
covid19clicked = 0
viewresultclicked = 0
analysis_mode = 0
sorted_contours1 = list(range(48))
temp_label = 0
name = "/"
entry_num = 0
wait = 0
pos_result = list(range(48))
t2_tmp= list(range(48))
path0 = "/"
path1 = "/"
path2 = "/"
path3 = "/"
path4 = "/"
path5 = "/"
filename = ""
importfilename = ""
excel_file = ""
id_list = list(range(48))
covid19_createclicked = 0
samples = 0
covid19dir_old = ""
div = list(range(48))
start_point = (0,0)
end_point = (0,0)
rsfile='/'
idfile='/'
test_list = list(range(48))
warning_value = 0
password = '123456789'
thr_set = 18
errors=0
detailclick = 0
readerror = 0

fr = open("/home/pi/Spotcheck/check.txt","r")
code = (fr.readline()).strip()
fr1 = open("/home/pi/Spotcheck/coordinates1.txt","r")
x1 = int(fr1.readline())
y1 = int(fr1.readline())
x2 = int(fr1.readline())
y2 = int(fr1.readline())
fr2 = open("/home/pi/Spotcheck/.server.txt","r")
server_on = int(fr2.readline())
ftp_ip = fr2.readline().strip('\n')
ftp_user = fr2.readline().strip('\n')
ftp_password = fr2.readline().strip('\n')
ftp_folder = fr2.readline().strip('\n')

hs = list(range(48))
tl = list(range(48))
workbook = openpyxl.load_workbook('/home/pi/Spotcheck/coefficient.xlsx')
sheet = workbook.active
for i in range(0,48):
    if(i<6):
        pos = str(chr(65+i+1)) + "2"
        pos1 = str(chr(65+i+1)) + "11"
    if(i>=6 and i<12):
        pos = str(chr(65+i-5)) + "3"
        pos1 = str(chr(65+i-5)) + "12"
    if(i>=12 and i<18):
        pos = str(chr(65+i-11)) + "4"
        pos1 = str(chr(65+i-11)) + "13"
    if(i>=18 and i<24):
        pos = str(chr(65+i-17)) + "5"
        pos1 = str(chr(65+i-17)) + "14"
    if(i>=24 and i<30):
        pos = str(chr(65+i-23)) + "6"
        pos1 = str(chr(65+i-23)) + "15"
    if(i>=30 and i<36):
        pos = str(chr(65+i-29)) + "7"
        pos1 = str(chr(65+i-29)) + "16"
    if(i>=36 and i<42):
        pos = str(chr(65+i-35)) + "8"
        pos1 = str(chr(65+i-35)) + "17"
    if(i>=42):
        pos = str(chr(65+i-41)) + "9"
        pos1 = str(chr(65+i-41)) + "18"
    hs[i] = float(sheet[pos].value)
    tl[i] = float(sheet[pos1].value)

fr3 = open("/var/tmp/.admin.txt","r")
start_trial = int(fr3.readline())
print("start_trial: ", start_trial)

fr7 = open("/home/pi/Spotcheck/parameters.txt")
average_min = float(fr7.readline())
average_max = float(fr7.readline())
hs_ct1 =  float(fr7.readline())
hs_ct2 =  float(fr7.readline())
a = float(fr7.readline())
b = float(fr7.readline())

if not os.path.exists('/home/pi/Spotcheck Ket Qua'):
    f = os.path.join("/home/pi/", "Spotcheck Ket Qua")
    os.mkdir(f)
if not os.path.exists('/home/pi/Desktop/Spotcheck ID'):
    f = os.path.join("/home/pi/Desktop", "Spotcheck ID")
    os.mkdir(f)
if not os.path.exists('/home/pi/Desktop/Spotcheck ID/Spotcheck ID - Old'):
    f = os.path.join("/home/pi/Desktop/Spotcheck ID/", "Spotcheck ID - Old")
    os.mkdir(f)
if not os.path.exists('/home/pi/Desktop/Ket Qua Phan Tich'):
    f = os.path.join("/home/pi/Desktop/", "Ket Qua Phan Tich")
    os.mkdir(f)
if not os.path.exists('/home/pi/Desktop/Ket Qua Phan Tich/Probe'):
    f = os.path.join("/home/pi/Desktop/Ket Qua Phan Tich/", "Probe")
    os.mkdir(f)
if not os.path.exists('/home/pi/Spotcheck Ket Qua/Probe'):
    f = os.path.join("/home/pi/Spotcheck Ket Qua/", "Probe")
    os.mkdir(f)

########################################################### GLOBAL VARIABLE - END ##################################################################

################################################################# TRIAL _ START ####################################################################
def trial():
    old_day = int(fr3.readline())
    old_month = int(fr3.readline())
    old_year = int(fr3.readline())
    limit = int(fr3.readline())
    print("Ng??y b???t ?????u:", old_day, old_month, old_year)

    today = datetime.now()
    new_day = today.day
    new_month = today.month
    new_year = today.year
    print("Ng??y hi???n t???i:", new_day, new_month, new_year)

    nam = new_year - old_year
    if(new_month < old_month):
        thang = new_month + 12 - old_month
        nam = nam - 1
    else:
        thang = new_month - old_month
    if(new_day < old_day):
        ngay = new_day + 30 - old_day
        thang = thang - 1
    else:
        ngay = new_day - old_day
    songay = ngay + thang*30 + nam*365
    print("Th???i gian d??ng th??? c??n l???i:", limit-songay)
    if(songay >= limit):
        trial_labelframe = LabelFrame(root, bg='white', width=800, height=600)
        trial_labelframe.place(x=0,y=0)

        logo_img = Image.open('/home/pi/Spotcheck/logo.png')
        logo_width, logo_height = logo_img.size
        scale_percent = 50
        width = int(logo_width * scale_percent / 100)
        height = int(logo_height * scale_percent / 100)
        display_img = logo_img.resize((width,height))
        image_select = ImageTk.PhotoImage(display_img)
        logo_label = Label(trial_labelframe, bg='white',image=image_select)
        logo_label.image = image_select
        logo_label.place(x=5,y=5)

        def active_click(event = None):
            code = activecode_entry.get()
            if(code==""):
                msg = messagebox.showwarning(" ","B???n ch??a nh???p m?? k??ch ho???t !")
            else:
                if(code!=password):
                    msg = messagebox.showerror(" ","M?? k??ch ho???t kh??ng ????ng !")
                if(code==password):
                    msg = messagebox.showinfo(" ","K??ch ho???t th??nh c??ng !")
                    f1=open("/var/tmp/.admin.txt",'w')
                    f1.writelines("0")
                    mainscreen()

        trial_label = Label(trial_labelframe, bg='white',fg="red", text="Th???i gian d??ng th??? ???? h???t\nVui l??ng nh???p m?? k??ch ho???t ????? ti???p t???c s??? d???ng !", font=("Courier",18,"bold"))
        trial_label.place(x=62,y=85)
        contact_label = Label(trial_labelframe, bg='white', text="Li??n h??? nh?? cung c???p ????? nh???n m?? k??ch ho???t:", font=("Courier",12,"bold"))
        contact_label.place(x=73,y=435)
        mail_label = Label(trial_labelframe, bg='white', fg='blue',text="cskh@phusabiochem.com", font=("Courier",12,"bold"))
        mail_label.place(x=503,y=435)
        activecode_entry = Entry(trial_labelframe, width=27, bg='white', font=("Courier",14,"bold"))
        activecode_entry.place(x=246,y=215)
        activecode_entry.bind("<Return>", active_click)
        code_label = Label(trial_labelframe, bg='white', text="M?? k??ch ho???t:", font=("Courier",14,"bold"))
        code_label.place(x=244,y=189)

        key_img = Image.open('/home/pi/Spotcheck/key.png')
        logo_width, logo_height = key_img.size
        scale_percent = 5
        width = int(logo_width * scale_percent / 100)
        height = int(logo_height * scale_percent / 100)
        display_img = key_img.resize((width,height))
        image_select = ImageTk.PhotoImage(display_img)
        logo_label = Label(trial_labelframe, bg='white',image=image_select)
        logo_label.image = image_select
        logo_label.place(x=726,y=85)

        active_button = Button(trial_labelframe, bg="lavender", font=("Courier",11,'bold'), text="X??c nh???n", height=3, width=10, borderwidth=0, command=active_click)
        active_button.place(x=340,y=260)
    else:
        mainscreen()
################################################################ TRIAL - END #######################################################################

########################################################## MAIN WINDOW INIT - START ################################################################
root = Tk()
root.title(" ")
root.geometry('1024x600')
root.configure(background = "white")
root.attributes('-fullscreen', True)
root.resizable(False,False)
def disable_event():
    pass
root.protocol("WM_DELETE_WINDOW", disable_event)
s = ttk.Style()
s.theme_use('clam')

if(code=='1111'):
    msgbox = messagebox.showerror(" ","H??? th???ng l???i, vui l??ng li??n h??? v???i nh?? cung c???p !")
    if(msgbox=='ok'):
        root.destroy()

########################################################### MAIN WINDOW INIT - END #################################################################

########################################################### RESOURCE PATH - START ##################################################################
def resoure_path(relative_path):
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)
############################################################ RESOURCE PATH - END ###################################################################

############################################################ CAMERA INIT - START ###################################################################
def camera_capture(output):
    global stop_click
    stop_click = 0
    camera = PiCamera(framerate=Fraction(1,6), sensor_mode=3)
    camera.rotation = 180
    camera.iso = 200
    sleep(2)
    camera.shutter_speed = 6000000
    camera.exposure_mode = 'off'
    camera.capture(output)
    camera.close()
############################################################# CAMERA INIT - END ####################################################################

############################################################ SERIAL INIT - START ###################################################################
ser = serial.Serial(
    port = '/dev/serial0',
    baudrate = 115200,
    parity = serial.PARITY_NONE,
    stopbits = serial.STOPBITS_ONE,
    bytesize = serial.EIGHTBITS,
    timeout = 1
)
############################################################# SERIAL INIT - END ####################################################################

######################################################### SORTING CONTOURS - START #################################################################
def sorting_y(contour):
    rect_y = cv2.boundingRect(contour)
    return rect_y[1]
def sorting_x(contour):
    rect_x = cv2.boundingRect(contour)
    return rect_x[0]
def sorting_xy(contour):
    rect_xy = cv2.boundingRect(contour)
    return math.sqrt(math.pow(rect_xy[0],2) + math.pow(rect_xy[1],2))
########################################################## SORTING CONTOURS - END ##################################################################

########################################################## IMAGE ANALYSIS - START ##################################################################
def process_image(image_name, start_point=(x1,y1), end_point=(x2,y2)):
    image = cv2.imread(image_name)
    blur_img = cv2.GaussianBlur(image.copy(), (35,35), 0)
    gray_img = cv2.cvtColor(blur_img, cv2.COLOR_BGR2GRAY)
    thresh, binary_img = cv2.threshold(gray_img.copy(), 30, maxval=255, type=cv2.THRESH_BINARY)
    contours, hierarchy = cv2.findContours(binary_img, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    print("Number of contours: " + str(len(contours)))

    contours.sort(key=lambda data:sorting_xy(data))

    contour_img = np.zeros_like(gray_img)
    contour_img = cv2.rectangle(contour_img, start_point, end_point, (255,255,255), -1)
    rect_w = end_point[0] - start_point[0]
    rect_h = end_point[1] - start_point[1]
    cell_w = round(rect_w/6)
    cell_h = round(rect_h/8)
    for i in range(1,6):
        contour_img = cv2.line(contour_img, (start_point[0]+i*cell_w,start_point[1]), (start_point[0]+i*cell_w,end_point[1]),(0,0,0), 4)
    for i in range(1,8):
        contour_img = cv2.line(contour_img, (start_point[0],start_point[1]+i*cell_h), (end_point[0],start_point[1]+i*cell_h),(0,0,0), 4)

    thresh1 , binary1_img = cv2.threshold(contour_img, 250, maxval=255, type=cv2.THRESH_BINARY)
    contours1, hierarchy1 = cv2.findContours(binary1_img, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_NONE)

    contours1.sort(key=lambda data:sorting_y(data))
    contours1_h1 = contours1[0:6]
    contours1_h2 = contours1[6:12]
    contours1_h3 = contours1[12:18]
    contours1_h4 = contours1[18:24]
    contours1_h5 = contours1[24:30]
    contours1_h6 = contours1[30:36]
    contours1_h7 = contours1[36:42]
    contours1_h8 = contours1[42:48]
    contours1_h1.sort(key=lambda data:sorting_x(data))
    contours1_h2.sort(key=lambda data:sorting_x(data))
    contours1_h3.sort(key=lambda data:sorting_x(data))
    contours1_h4.sort(key=lambda data:sorting_x(data))
    contours1_h5.sort(key=lambda data:sorting_x(data))
    contours1_h6.sort(key=lambda data:sorting_x(data))
    contours1_h7.sort(key=lambda data:sorting_x(data))
    contours1_h8.sort(key=lambda data:sorting_x(data))

    sorted_contours1 = contours1_h1 + contours1_h2 + contours1_h3 + contours1_h4 + contours1_h5 + contours1_h6 + contours1_h7 + contours1_h8

    list_intensities = []
    sum_intensities = []
    result_list = list(range(48))
    area = list(range(48))

#Gray
#     tmp_list = list(range(48))
#     blur1_img = cv2.GaussianBlur(image.copy(), (25,25), 0)
#     grayprocess_img = cv2.cvtColor(blur1_img, cv2.COLOR_BGR2GRAY)
#     #cv2.imwrite("mau.jpg",grayprocess_img)
#     for i in range(len(sorted_contours1)):
#         cimg = np.zeros_like(gray_img)
#         cv2.drawContours(cimg, sorted_contours1, i, color = 255, thickness = -1)
#         pts = np.where(cimg == 255)
#         list_intensities.append(grayprocess_img[pts[0], pts[1]])
#         list_intensities[i].sort()
#         #print("list_intensities",str(i),":",list_intensities[i])
#         #print("value", str(i), " : ", list_intensities[i][len(list_intensities[i])-1])
#         sum_intensities.append(sum(list_intensities[i][len(list_intensities[i])-280:]))
#         #sum_intensities.append(sum(list_intensities[i][len(list_intensities[i])-240:]))
#         area[i]= cv2.contourArea(sorted_contours1[i])
#         #result_list[i] = sum_intensities[i]
#         tmp_list[i] = sum_intensities[i]/1000
#         #result_list[i] = round(tmp_list[i])
#         #result_list[i] = round(round(tmp_list[i],1)*1.5)
#         result_list[i] = round(tmp_list[i],1)

#BGR
    blur1_img = cv2.GaussianBlur(image.copy(), (25,25), 0)
    tmp_list = list(range(48))
    list_bgrvalue = []
    list_index = list(range(48))
    for i in range(len(sorted_contours1)):
        list_index[i] = []
        cimg = np.zeros_like(gray_img)
        cv2.drawContours(cimg, sorted_contours1, i, color = 255, thickness = -1)
        pts = np.where(cimg == 255)
        list_bgrvalue.append(blur1_img[pts[0], pts[1]])
        for j in range(len(list_bgrvalue[i])):
             list_index[i].append(round((list_bgrvalue[i][j][1]*3 + list_bgrvalue[i][j][2])))
        list_index[i].sort()
        list_intensities.append(sum(list_index[i][len(list_index[i])-250:]))
        area[i]= cv2.contourArea(sorted_contours1[i])
        tmp_list[i] = list_intensities[i]/1000
        result_list[i] = round(tmp_list[i],1)

#HSV
#     tmp_list = list(range(48))
#     #blur1_img = cv2.fastNlMeansDenoisingColored(image.copy(),None,9,9,7,19)
#     #cv2.imwrite("mau1.jpg",blur1_img)
#     blur1_img = cv2.GaussianBlur(image.copy(), (3,3), 0)
#     cv2.imwrite("mau.jpg",blur_img)
#     hsv_img = cv2.cvtColor(blur1_img, cv2.COLOR_BGR2HSV)
#     list_hsvvalue = []
#     list_index = list(range(48))
#     for i in range(len(sorted_contours1)):
#         list_index[i] = []
#         cimg = np.zeros_like(gray_img)
#         cv2.drawContours(cimg, sorted_contours1, i, color = 255, thickness = -1)
#         pts = np.where(cimg == 255)
#         list_hsvvalue.append(hsv_img[pts[0], pts[1]])
#         for j in range(len(list_hsvvalue[i])):
#             list_index[i].append(list_hsvvalue[i][j][2])
#         list_index[i].sort()
#         #print(len(list_index[i]))
#         list_intensities.append(sum(list_index[i][len(list_index[i])-250:]))
#         #area[i]= cv2.contourArea(sorted_contours1[i])
#         result_list[i] = list_intensities[i]
#         tmp_list[i] = list_intensities[i]/1000
#         result_list[i] = round(tmp_list[i])

#Nh??n h??? s???
    global hs
    for i in range(1,7):
        result_list[6*i+1]=round(result_list[6*i+1]*(1-(0.02*round(result_list[6*i]/70) + 0.02*round(result_list[6*i+2]/70) + 0.02*round(result_list[6*i-5]/70) + 0.02*round(result_list[6*i+7]/70) + 0.003*round(result_list[6*i-6]/70) + 0.003*round(result_list[6*i-4]/70) + 0.003*round(result_list[6*i+6]/70) + 0.003*round(result_list[6*i+8]/70)+ 0.006*round(result_list[6*i+3]/70))),1)
        result_list[6*i+2]=round(result_list[6*i+2]*(1-(0.02*round(result_list[6*i+1]/70) + 0.02*round(result_list[6*i+3]/70) + 0.02*round(result_list[6*i-4]/70) + 0.02*round(result_list[6*i+8]/70) + 0.003*round(result_list[6*i-5]/70) + 0.003*round(result_list[6*i-3]/70) + 0.003*round(result_list[6*i+7]/70) + 0.003*round(result_list[6*i+9]/70)+ 0.006*round(result_list[6*i+4]/70)+ 0.006*round(result_list[6*i]/70))),1)
        result_list[6*i+3]=round(result_list[6*i+3]*(1-(0.02*round(result_list[6*i+2]/70) + 0.02*round(result_list[6*i+4]/70) + 0.02*round(result_list[6*i-3]/70) + 0.02*round(result_list[6*i+9]/70) + 0.003*round(result_list[6*i-4]/70) + 0.003*round(result_list[6*i-2]/70) + 0.003*round(result_list[6*i+8]/70) + 0.003*round(result_list[6*i+10]/70)+ 0.006*round(result_list[6*i+5]/70)+ 0.006*round(result_list[6*i+1]/70))),1)
        result_list[6*i+4]=round(result_list[6*i+4]*(1-(0.02*round(result_list[6*i+3]/70) + 0.02*round(result_list[6*i+5]/70) + 0.02*round(result_list[6*i-2]/70) + 0.02*round(result_list[6*i+10]/70) + 0.003*round(result_list[6*i-3]/70) + 0.003*round(result_list[6*i-1]/70) + 0.003*round(result_list[6*i+9]/70) + 0.003*round(result_list[6*i+11]/70)+ 0.006*round(result_list[6*i+2]/70))),1)

        result_list[6*i]=round(result_list[6*i]*(1-(0.02*round(result_list[6*i+1]/70) + 0.015*round(result_list[6*i-6]/70) + 0.015*round(result_list[6*i+6]/70) + 0.003*round(result_list[6*i-5]/70) + 0.003*round(result_list[6*i+7]/70)+0.006*round(result_list[6*i+2]/70))),1)
        result_list[6*i+5]=round(result_list[6*i+5]*(1-(0.02*round(result_list[6*i+4]/70) + 0.015*round(result_list[6*i-1]/70) + 0.015*round(result_list[6*i+11]/70) + 0.003*round(result_list[6*i-2]/70) + 0.003*round(result_list[6*i+10]/70)+0.006*round(result_list[6*i+3]/70))),1)

    for i in range(1,5):
        result_list[i]=round(result_list[i]*(1-(0.02*round(result_list[i-1]/70) + 0.02*round(result_list[i+1]/70) + 0.015*round(result_list[i+6]/70)+ 0.003*round(result_list[i+5]/70) + 0.003*round(result_list[i+7]/70))),1)
        result_list[i+42]=round(result_list[i+42]*(1-(0.02*round(result_list[i+41]/70) + 0.02*round(result_list[i+43]/70) + 0.015*round(result_list[i+36]/70)+ 0.003*round(result_list[i+35]/70) + 0.003*round(result_list[i+37]/70))),1)

    result_list[0]=round(result_list[0]*(1-(0.015*round(result_list[1]/70) + 0.015*round(result_list[6]/70))),1)
    result_list[5]=round(result_list[5]*(1-(0.015*round(result_list[4]/70) + 0.015*round(result_list[11]/70))),1)
    result_list[42]=round(result_list[42]*(1-(0.015*round(result_list[43]/70) + 0.015*round(result_list[36]/70))),1)
    result_list[47]=round(result_list[47]*(1-(0.015*round(result_list[46]/70) + 0.015*round(result_list[41]/70))),1)

    if(analysis_mode == 1):
        for i in range(len(sorted_contours1)):
            if(id_list[i]=='N' or id_list[i]=='Negative' or id_list[i]=='n' or id_list[i]=='negative' or id_list[i]=='NEGC' or id_list[i]=='NEGATIVE'):
                result_list[i] = round(result_list[i]*0.85,1)
            elif(id_list[i]=='P' or id_list[i]=='Positive' or id_list[i]=='p' or id_list[i]=='positive' or id_list[i]=='POSC' or id_list[i]=='POSITIVE'):
                result_list[i] = round(result_list[i]*1.15,1)
            else:
                result_list[i] = round(result_list[i]*hs[i],1)

#     for i in range(len(sorted_contours1)):
#         if(i==0):
#             result_list[i] = round(result_list[i]*1.03,1)
#         if(i==47):
#             result_list[i] = round(result_list[i]*0.92,1)
#         if(i==37 or i==38 or i==39 or i==40 or
#            i==43 or i==44 or i==45 or
#            i==10 or i==16 or i==22 or i==28 or i==34):
#             result_list[i] = round(result_list[i]*0.98,1)
#         if(i==36 or i==42 or i==46 or i==4 or
#            i==5 or i==11 or i==17 or i==23 or i==29 or i==35 or i==41):
#             result_list[i] = round(result_list[i]*0.96,1)

    for i in range(len(sorted_contours1)):
        if(result_list[i]>99):
            result_list[i]=99

    for i in range(len(sorted_contours1)):
        if ((i!=0) and ((i+1)%6==0)):
            print('%.1f'%(result_list[i]))
        else:
            print('%.1f'%(result_list[i]), end = ' | ')

    blurori_img = cv2.GaussianBlur(image.copy(), (25,25), 0)
    for i in range(len(sorted_contours1)):
        if(id_list[i]=='N/A'):
            cv2.drawContours(blurori_img, sorted_contours1, i, (0,0,0), thickness = -1)
        else:
            if(result_list[i] <= 0.8*float(thr_set)*1.05 and result_list[i] >= 0.8*float(thr_set)*0.95):
                cv2.drawContours(blurori_img, sorted_contours1, i, (0,255,0), thickness = 2)
            else:
                cv2.drawContours(blurori_img, sorted_contours1, i, (0,0,255), thickness = 2)
    return (result_list, blurori_img)
########################################################### IMAGE ANALYSIS - END ###################################################################

############################################################ MAIN SCREEN - START ###################################################################
def mainscreen():
    buttonFont = font.Font(family='Helvetica', size=10, weight='bold')
    global mainscreen_labelframe
    mainscreen_labelframe = LabelFrame(root, bg='white', width=800, height=480)
    mainscreen_labelframe.place(x=0,y=0)
    sidebar_labelframe = LabelFrame(mainscreen_labelframe, font=("Courier",15,'bold'), bg='dodger blue', width=172, height=476)
    sidebar_labelframe.place(x=0,y=0)

    global analysis_mode
    analysis_mode = 0

    def home_click():
        try:
            subprocess.Popen(['killall','florence'])
        except:
            pass
        root.attributes('-fullscreen', True)

        home_canvas['bg'] = 'white'
        covid19_canvas['bg'] = 'dodger blue'
        viewresult_canvas['bg'] = 'dodger blue'
        setid_canvas['bg'] = 'dodger blue'
        config_canvas['bg'] = 'dodger blue'
        power_canvas['bg'] = 'dodger blue'

        global covid19clicked
        covid19clicked = 0

        global covid19_createclicked
        covid19_createclicked = 0

        homemc_labelframe = LabelFrame(mainscreen_labelframe, bg='white', width=625, height=476)
        homemc_labelframe.place(x=172,y=0)

        top_labelframe = LabelFrame(homemc_labelframe, bg='dodger blue', width=619, height=37)
        top_labelframe.place(x=1,y=1)

        hometitle_label = Label(top_labelframe, bg='dodger blue', font=("Courier", 11, 'bold'), text="TRANG CH???")
        hometitle_label.place(x=260,y=7)

        probe_label = Label(homemc_labelframe, bg='white', fg = "grey80", font=("Courier", 10, 'bold'), text="Probe version")
        probe_label.place(x=5,y=449)

        logo_img = Image.open('/home/pi/Spotcheck/logo.png')
        logo_width, logo_height = logo_img.size
        scale_percent = 30
        width = int(logo_width * scale_percent / 100)
        height = int(logo_height * scale_percent / 100)
        display_img = logo_img.resize((width,height))
        image_select = ImageTk.PhotoImage(display_img)
        logo_label = Label(mainscreen_labelframe, bg='white',image=image_select)
        logo_label.image = image_select
        logo_label.place(x=650,y=442)

        home_img = Image.open('/home/pi/Spotcheck/home.png')
        home_width, home_height = home_img.size
        scale_percent = 100
        width = int(home_width * scale_percent / 100)
        height = int(home_height * scale_percent / 100)
        display_img = home_img.resize((width,height))
        image_select = ImageTk.PhotoImage(display_img)
        logo_label = Label(mainscreen_labelframe, bg='white',image=image_select)
        logo_label.image = image_select
        logo_label.place(x=231,y=115)

        readerror_label = Label(mainscreen_labelframe, bg='red', text="H??? TH???NG L???I!", font=("Courier",14,"bold"))
        if(readerror==1):
            readerror_label.place(x=642,y=9)
            root.update_idletasks()
        else:
            readerror_label.place(x=1000,y=1000)
            root.update_idletasks()

    def covid19_click():
        try:
                subprocess.Popen(['killall','florence'])
        except:
            pass
        root.attributes('-fullscreen', True)

        home_canvas['bg'] = 'dodger blue'
        covid19_canvas['bg'] = 'white'
        viewresult_canvas['bg'] = 'dodger blue'
        setid_canvas['bg'] = 'dodger blue'
        config_canvas['bg'] = 'dodger blue'
        power_canvas['bg'] = 'dodger blue'

        global covid19clicked
        covid19clicked = 1

        global spotcheck_createclicked, tb_createclicked, shrimp_createclicked
        spotcheck_createclicked = 0

        covid19mc_labelframe = LabelFrame(mainscreen_labelframe, bg='white', width=624, height=478)
        covid19mc_labelframe.place(x=172,y=0)

        top_labelframe = LabelFrame(covid19mc_labelframe, bg='dodger blue', width=619, height=37)
        top_labelframe.place(x=1,y=1)

        covid19title_label = Label(top_labelframe, bg='dodger blue', font=("Courier", 11, 'bold'), text="PH??N T??CH")
        covid19title_label.place(x=260,y=7)

        enterframe_labelframe = LabelFrame(covid19mc_labelframe, bg='white', width=609, height=175)
        enterframe_labelframe.place(x=5,y=46)

        folder1name_label = Label(enterframe_labelframe, bg='yellow2',width=75, height=1)
        folder1name_label.place(x=0,y=1)

        foldername_label = Label(enterframe_labelframe, bg='white',text='T???p m???u x??t nghi???m', fg='black', font=("Courier",12,'bold'))
        foldername_label.place(x=75,y=37)

        # import_label = Label(enterframe_labelframe, bg='white',text='ID file', fg='black', font=("Courier",11,'bold'))
        # import_label.place(x=68,y=80)

        file_label = Label(enterframe_labelframe, bg='white', fg='grey25', font=("Courier",13,'bold'))
        file_label.place(x=213,y=87)

        global covid19_createclicked
        global covid19dir_old
        global importfilename
        directory = strftime("COVID19 %y-%m-%d %H.%M.%S")
        if(covid19_createclicked == 0):
            directory_label = Label(enterframe_labelframe, font=("Courier",10,'bold'), bg='yellow2', text=directory)
            directory_label.place(x=190,y=1)
            covid19dir_old = directory
        else:
            directory_label = Label(enterframe_labelframe, font=("Courier",10,'bold'), bg='yellow2', text=covid19dir_old)
            directory_label.place(x=190,y=1)
            file_label['text'] = importfilename

        def thread():
            th1 = Thread(target = create_click)
            th1.start()
        def create_click(event=None):
            global covid19_createclicked
            covid19_createclicked = 1
            global foldername
            #foldername = str(file_label['text'])
            name = strftime(importfilename)
            global path0
            global covid19dir_old
            print("covid19dir_old:",covid19dir_old)
            path0 = os.path.join("/home/pi/Spotcheck Ket Qua/Probe/", covid19dir_old +" "+ name +"/")

            if(file_label['text']==""):
                msgbox = messagebox.showwarning(" ","B???n ch??a t???i t???p l??n !" )
            else:
                if os.path.exists(path0):
                    msg = messagebox.askquestion("Th?? m???c ???? t???n t???i", "B???n c?? mu???n thay th??? th?? m???c c?? ?")
                    if(msg=='yes'):
                        shutil.rmtree(path0)
                        os.mkdir(path0)
                        global path1
                        path1 = os.path.join(path0,"???nh ch???p")
                        os.mkdir(path1)
                        global path2
                        path2 = os.path.join(path0,"???nh x??? l??")
                        os.mkdir(path2)
                        global path3
                        path3 = os.path.join(path0,"B???ng k???t qu???")
                        os.mkdir(path3)
                        global path4
                        path4 = os.path.join(path0,"???nh nguy??n m???u")
                        os.mkdir(path4)
                        global path5
                        path5 = os.path.join(path0,"Ch????ng tr??nh nhi???t")
                        os.mkdir(path5)
                        mainscreen_labelframe.place_forget()
                        scanposition()
                else:
                    os.mkdir(path0)
                    path1 = os.path.join(path0,"???nh ch???p")
                    os.mkdir(path1)
                    path2 = os.path.join(path0,"???nh x??? l??")
                    os.mkdir(path2)
                    path3 = os.path.join(path0,"B???ng k???t qu???")
                    os.mkdir(path3)
                    path4 = os.path.join(path0,"???nh nguy??n m???u")
                    os.mkdir(path4)
                    path5 = os.path.join(path0,"Ch????ng tr??nh nhi???t")
                    os.mkdir(path5)

                    mainscreen_labelframe.place_forget()
                    scanposition()

        def import_click():
            if(server_on==1):
                try:
                    ftp = FTP(ftp_ip, ftp_user, ftp_password, timeout=15)
                    ftp.cwd(ftp_folder + 'UnProcessed_Data')
                    ftpfiles = ftp.nlst()
                    for ftpfile in ftpfiles:
                        if(os.path.exists("/home/pi/Desktop/Spotcheck ID/" + ftpfile)):
                            pass
                        elif(os.path.exists("/home/pi/Desktop/Spotcheck ID/Spotcheck ID - Old/" + ftpfile)):
                            pass
                        else:
                            localfolder = os.path.join('/home/pi/Desktop/Spotcheck ID/', ftpfile)
                            file = open(localfolder,'wb')
                            ftp.retrbinary('RETR ' + ftpfile, file.write)
                            file.close()
                            print(ftpfile, "download done!")
                    ftp.quit()
                except Exception as e :
                    error = messagebox.showwarning("C?? l???i x???y ra khi ?????ng b??? server !",str(e))
                    if(error=='ok'):
                        pass
            file = filedialog.askopenfile(initialdir='/home/pi/Desktop/Spotcheck ID/', mode='r', filetypes=[('Excel file','*.xlsm *.xlsx *.xls')])
            global importfilename
            global filename
            filename = file.name
            global excel_file
            if file is not None:
                a=0
                for i in range(len(filename)):
                    if(filename[i]=='/'):
                        a=i+1
                importfilename = filename[a:(len(filename)-5)]
                if (os.path.exists("/home/pi/Desktop/Ket Qua Phan Tich/Probe/" + importfilename + ".xlsm")):
                    messagebox.showwarning("","T???p v???a ch???n ???? ???????c s??? d???ng !")
                    create_button['state']='disabled'
                    create_button['bg']='grey75'
                    file_label['text'] = ""
                    #import_click()
                else:
                    excel_file = filename[a:len(filename)]
                    if(len(importfilename)>=15):
                        file_label['text'] = importfilename[0:15] + '...'
                    else:
                        file_label['text'] = importfilename

                    workbook = openpyxl.load_workbook(filename)
                    sheet = workbook.active
                    # for i in range(0,48):
                    #     if(i<6):
                    #         pos = str(chr(65+i+1)) + "2"
                    #     if(i>=6 and i<12):
                    #         pos = str(chr(65+i-5)) + "3"
                    #     if(i>=12 and i<18):
                    #         pos = str(chr(65+i-11)) + "4"
                    #     if(i>=18 and i<24):
                    #         pos = str(chr(65+i-17)) + "5"
                    #     if(i>=24 and i<30):
                    #         pos = str(chr(65+i-23)) + "6"
                    #     if(i>=30 and i<36):
                    #         pos = str(chr(65+i-29)) + "7"
                    #     if(i>=36 and i<42):
                    #         pos = str(chr(65+i-35)) + "8"
                    #     if(i>=42):
                    #         pos = str(chr(65+i-41)) + "9"

                    tmp_list = list(range(48))
                    for i in range(0,48):
                        pos = "B" + str(i+12)
                        tmp_list[i] = sheet[pos].value
                        if(i==0):
                            id_list[0] = tmp_list[i]
                        if(i==1):
                            id_list[6] = tmp_list[i]
                        if(i==2):
                            id_list[12] = tmp_list[i]
                        if(i==3):
                            id_list[18] = tmp_list[i]
                        if(i==4):
                            id_list[24] = tmp_list[i]
                        if(i==5):
                            id_list[30] = tmp_list[i]
                        if(i==6):
                            id_list[36] = tmp_list[i]
                        if(i==7):
                            id_list[42] = tmp_list[i]
                        if(i==8):
                            id_list[1] = tmp_list[i]
                        if(i==9):
                            id_list[7] = tmp_list[i]
                        if(i==10):
                            id_list[13] = tmp_list[i]
                        if(i==11):
                            id_list[19] = tmp_list[i]
                        if(i==12):
                            id_list[25] = tmp_list[i]
                        if(i==13):
                            id_list[31] = tmp_list[i]
                        if(i==14):
                            id_list[37] = tmp_list[i]
                        if(i==15):
                            id_list[43] = tmp_list[i]
                        if(i==16):
                            id_list[2] = tmp_list[i]
                        if(i==17):
                            id_list[8] = tmp_list[i]
                        if(i==18):
                            id_list[14] = tmp_list[i]
                        if(i==19):
                            id_list[20] = tmp_list[i]
                        if(i==20):
                            id_list[26] = tmp_list[i]
                        if(i==21):
                            id_list[32] = tmp_list[i]
                        if(i==22):
                            id_list[38] = tmp_list[i]
                        if(i==23):
                            id_list[44] = tmp_list[i]
                        if(i==24):
                            id_list[3] = tmp_list[i]
                        if(i==25):
                            id_list[9] = tmp_list[i]
                        if(i==26):
                            id_list[15] = tmp_list[i]
                        if(i==27):
                            id_list[21] = tmp_list[i]
                        if(i==28):
                            id_list[27] = tmp_list[i]
                        if(i==29):
                            id_list[33] = tmp_list[i]
                        if(i==30):
                            id_list[39] = tmp_list[i]
                        if(i==31):
                            id_list[45] = tmp_list[i]
                        if(i==32):
                            id_list[4] = tmp_list[i]
                        if(i==33):
                            id_list[10] = tmp_list[i]
                        if(i==34):
                            id_list[16] = tmp_list[i]
                        if(i==35):
                            id_list[22] = tmp_list[i]
                        if(i==36):
                            id_list[28] = tmp_list[i]
                        if(i==37):
                            id_list[34] = tmp_list[i]
                        if(i==38):
                            id_list[40] = tmp_list[i]
                        if(i==39):
                            id_list[46] = tmp_list[i]
                        if(i==40):
                            id_list[5] = tmp_list[i]
                        if(i==41):
                            id_list[11] = tmp_list[i]
                        if(i==42):
                            id_list[17] = tmp_list[i]
                        if(i==43):
                            id_list[23] = tmp_list[i]
                        if(i==44):
                            id_list[29] = tmp_list[i]
                        if(i==45):
                            id_list[35] = tmp_list[i]
                        if(i==46):
                            id_list[41] = tmp_list[i]
                        if(i==47):
                            id_list[47] = tmp_list[i]

                    create_button['state']='normal'
                    create_button['bg']='lawn green'

        import_button = Button(enterframe_labelframe, font=("Courier",10,'bold'), bg="lavender", text="T???i l??n", height=3, width=10, borderwidth=0, command=import_click)
        import_button.place(x=78,y=69)

        if(file_label['text']==""):
            create_button = Button(enterframe_labelframe, font=("Courier",10,'bold'), bg="grey75", text="Ti???p theo", height=3, width=10, borderwidth=0, command=thread, state='disabled')
        else:
            create_button = Button(enterframe_labelframe, font=("Courier",10,'bold'), bg="lawn green", text="Ti???p theo", height=3, width=10, borderwidth=0, command=thread)
        create_button.place(x=403,y=69)

        if(warning_value==1):
            warning_label = Label(mainscreen_labelframe, bg='red',text='H??? th???ng ??ang t???n nhi???t, kh??ng ?????t m???u v??o l??c n??y !', font=("Courier", 13, 'bold'))
            warning_label.place(x=220,y=450)
        else:
            warning_label = Label(mainscreen_labelframe, bg='white', fg='white', text='H??? th???ng ??ang t???n nhi???t, kh??ng ?????t m???u v??o l??c n??y !', font=("Courier", 13, 'bold'))
            warning_label.place(x=220,y=450)

        readerror_label = Label(mainscreen_labelframe, bg='red', text="H??? TH???NG L???I!", font=("Courier",14,"bold"))
        if(readerror==1):
            readerror_label.place(x=642,y=9)
            root.update_idletasks()
        else:
            readerror_label.place(x=1000,y=1000)
            root.update_idletasks()

    def setid_click():
        try:
                subprocess.Popen(['killall','florence'])
        except:
            pass
        root.attributes('-fullscreen', True)

        home_canvas['bg'] = 'dodger blue'
        covid19_canvas['bg'] = 'dodger blue'
        viewresult_canvas['bg'] = 'dodger blue'
        setid_canvas['bg'] = 'white'
        config_canvas['bg'] = 'dodger blue'
        power_canvas['bg'] = 'dodger blue'

        setidmc_labelframe = LabelFrame(mainscreen_labelframe, bg='white', width=624, height=478)
        setidmc_labelframe.place(x=172,y=0)

        top_labelframe = LabelFrame(setidmc_labelframe, bg='dodger blue', width=619, height=37)
        top_labelframe.place(x=1,y=1)

        setidtitle_label = Label(top_labelframe, bg='dodger blue', font=("Courier", 11, 'bold'), text="T???P M???U X??T NGHI???M")
        setidtitle_label.place(x=215,y=7)

        # setid_labelframe = LabelFrame(setidmc_labelframe, bg='white', width=430, height=435)
#         setid_labelframe.place(x=95,y=15)

        # setid0_label = Label(setid_labelframe, bg='dodger blue', text=" V??? TR?? ?????T M???U", font=("Courier", 12,'bold'),width=42)
#         setid0_label.place(x=1,y=1)

        # s48_img = Image.open('/home/pi/Spotcheck/48well.JPG')
#         s48_width, s48_height = s48_img.size
#         scale_percent = 50
#         width = int(s48_width * scale_percent / 100)
#         height = int(s48_height * scale_percent / 100)
#         display_img = s48_img.resize((width,height))
#         image_select = ImageTk.PhotoImage(display_img)
#         setid_label = Label(setidmc_labelframe, bg='white',image=image_select)
#         setid_label.image = image_select
#         setid_label.place(x=99,y=40)

#         def ok_click():
#             global covid19_createclicked
#             covid19_createclicked = 0
#             setid()

#         ok_button = Button(setidmc_labelframe, fg='black', font=('Courier','13','bold'), bg="lavender", text="Ti???p theo", height=2, width=10, borderwidth=0, command=ok_click)
#         ok_button.place(x=150,y=355)


        setid2_labelframe = LabelFrame(setidmc_labelframe, bg='white', width=315, height=192)
        setid2_labelframe.place(x=305,y=38)

        idpos_label = Label(setid2_labelframe, bg='turquoise2', font=("Courier",15,"bold"))
        idpos_label.place(x=1,y=1)

        setidtable_labelframe = LabelFrame(setidmc_labelframe,bg='grey95', width=600, height=307)
        setidtable_labelframe.place(x=1,y=38 )

        def idpos_click(n):
            if(idpos_button[n]['bg'] != 'lawn green'):
                for k in range (0,48):
                    if(idpos_button[k]['bg'] != 'lawn green' and idpos_button[k]['bg'] != 'grey99'):
                        idpos_button[k]['bg'] = 'lavender'
                    else:
                        idpos_button[k]['bg'] = 'lawn green'
                idpos_button[n]['bg'] = 'turquoise2'
            else:
                for k in range (0,48):
                    if(idpos_button[k]['bg'] != 'lawn green' and idpos_button[k]['bg'] != 'grey99'):
                        idpos_button[k]['bg'] = 'lavender'
                    if(idpos_button[k]['bg'] == 'grey99'):
                        idpos_button[k]['bg'] = 'lawn green'
                idpos_button[n]['bg'] = 'grey99'

            def enter_entry(event):
                try:
                    subprocess.Popen(['killall','florence'])
                except:
                    pass
                root.attributes('-fullscreen', False)
                subprocess.Popen('florence',stdout=subprocess.PIPE, shell=True)
                subprocess.Popen('florence',stdout=subprocess.PIPE, shell=True)

            def ok_click(event=None):
                if(id_entry.get()==''):
                    idpos_button[n]['bg'] = 'lavender'
                    idpos_button[n]['text'] = '#'+str(n+1)
                    msgbox = messagebox.showwarning(" ","B???n ch??a nh???p ID !")
                else:
                    idpos_button[n]['text'] = id_entry.get()
                    idpos_button[n]['bg'] = 'lawn green'
                    try:
                        if(n==47):
                            idpos_click(0)
                        else:
                            idpos_click(n+1)
                    except:
                        idpos_click(0)

            id_entry = Entry(setid2_labelframe,width=25, font=('Courier',14))
            if(idpos_button[n]['bg'] == 'grey99'):
                id_entry.insert(0,idpos_button[n]['text'])
            #id_entry.bind("<Button-1>", enter_entry)
            id_entry.bind("<Return>", ok_click)
            id_entry.place(x=15,y=75)
            id_entry.focus_set()

            setid_label = Label(setid2_labelframe, text='Nh???p m???u x??t nghi???m', bg='white', font=("Courier",15,"bold"))
            setid_label.place(x=15,y=45)

            if(n<8):
                idpos_label['text'] = str(chr(65+n)) + '1'
            if(n>=8 and n<16):
                idpos_label['text'] = str(chr(65+n-8)) + '2'
            if(n>=16 and n<24):
                idpos_label['text'] = str(chr(65+n-16)) + '3'
            if(n>=24 and n<32):
                idpos_label['text'] = str(chr(65+n-24)) + '4'
            if(n>=32 and n<40):
                idpos_label['text'] = str(chr(65+n-32)) + '5'
            if(n>=40):
                idpos_label['text'] = str(chr(65+n-40)) + '6'

            ok_button = Button(setid2_labelframe, font=('Courier','10','bold'), bg="lavender", text="X??c nh???n", height=3, width=9, borderwidth=0, command=ok_click)
            ok_button.place(x=198,y=113)

        idpos_button = list(range(48))
        h=-1
        c=0
        for i in range(0,48):
            h+=1
            if(i%8==0 and i!=0):
                h=0
                c+=1
            idpos_button[i] = Button(setidtable_labelframe, bg='lavender', activebackground="white", justify='left', borderwidth=0, text='#'+str(i+1), width=2, height=2)
            idpos_button[i]['command'] = partial(idpos_click,i)
            idpos_button[i].grid(row=h,column=c,padx=4,pady=4)
            # if(i==46):
            #     idpos_button[i]['state']='disabled'
            #     idpos_button[i]['bg']= 'green'
            #     idpos_button[i]['text']= 'N'
            # if(i==47):
            #     idpos_button[i]['state']='disabled'
            #     idpos_button[i]['bg']= 'red'
            #     idpos_button[i]['text']= 'P'

        def cancel_click():
            msg = messagebox.askquestion("H???y", "B???n mu???n h???y m?? kh??ng l??u l???i t???p ?")
            if(msg=="yes"):
                setid_click()

        def save_click():
            workbook = load_workbook("/home/pi/Spotcheck/template.xlsm", keep_vba = True)
            sheet = workbook.active
            # for i in range(0,48):
            #     #pos = "C"+str(i+3)
            #     if(i<8):
            #         pos = 'B'+ str(i+2)
            #     if(i>=8 and i<16):
            #         pos = 'C'+ str(i-6)
            #     if(i>=16 and i<24):
            #         pos = 'D'+ str(i-14)
            #     if(i>=24 and i<32):
            #         pos = 'E'+ str(i-22)
            #     if(i>=32 and i<40):
            #         pos = 'F'+ str(i-30)
            #     if(i>=40):
            #         pos = 'G'+str(i-38)

            for i in range(0,48):
                pos = "B" + str(i+12)
                if(idpos_button[i]['bg']=='lawn green' or idpos_button[i]['bg']=='grey99'):
                    sheet[pos] = idpos_button[i]['text']
                else:
                    sheet[pos] = 'N/A'

            # sheet['B58']='NEGC'
            # sheet['B59']='POSC'

            msg = messagebox.askquestion("L??u ", "B???n c?? mu???n l??u t???p ?")
            if(msg=='yes'):
                f = filedialog.asksaveasfilename(initialdir='/home/pi/Desktop/Spotcheck ID/',defaultextension='.xlsx')
                if f is not None:
                    d=0
                    for i in range(len(f)):
                        if(f[i]=='/'):
                            d=i+1
                    filename = f[d:(len(f)-5)]
                    print(filename)
                    if(len(filename)<=30):
                        workbook.save(f)
                        try:
                            subprocess.Popen(['killall','florence'])
                        except:
                            pass
                        root.attributes('-fullscreen', True)

                        msg = messagebox.askquestion(' ','???? l??u xong!\nB???n c?? mu???n t???o t???p m???i ?')
                        if(msg=='yes'):
                            setid_click()
                        else:
                            covid19_click()

                    else:
                        messagebox.showerror("L???i", "T??n t???p kh??ng v?????t qu?? 30 k?? t??? !")

        def load_click():
            idfile = filedialog.askopenfilename(initialdir='/home/pi/Desktop/Spotcheck ID', filetypes=[('Excel file','*.xlsm *.xlsx *.xls')])
            if idfile is not None:
                if(idfile[len(idfile)-4:]=='xlsx' or idfile[len(idfile)-4:]=='xlsm' or idfile[len(idfile)-3:]=='xls'):
                    workbook = openpyxl.load_workbook(idfile)
                    sheet = workbook.active
                    idfile_list = list(range(48))

                    # for i in range(0,48):
                    #     if(i<8):
                    #         pos = 'B'+ str(i+2)
                    #     if(i>=8 and i<16):
                    #         pos = 'C'+ str(i-6)
                    #     if(i>=16 and i<24):
                    #         pos = 'D'+ str(i-14)
                    #     if(i>=24 and i<32):
                    #         pos = 'E'+ str(i-22)
                    #     if(i>=32 and i<40):
                    #         pos = 'F'+ str(i-30)
                    #     if(i>=40):
                    #         pos = 'G'+str(i-38)

                    for i in range(0,48):
                        pos = 'B' + str(i+12)
                        idfile_list[i] = sheet[pos].value
                        idpos_button[i]['text'] = idfile_list[i]
                        if(idpos_button[i]['text']!='N/A'):
                            idpos_button[i]['bg']='lawn green'
                        else:
                            idpos_button[i]['bg']='lavender'
                        # if(i==46):
                        #     idpos_button[i]['bg']= 'green'
                        # if(i==47):
                        #     idpos_button[i]['bg']= 'red'
            else:
                pass
        def keyboard_click():
            if(keyboard_button['bg']=='grey85'):
                keyboard_button['bg']='lawn green'
                try:
                    subprocess.Popen(['killall','florence'])
                except:
                    pass
                root.attributes('-fullscreen', False)
                subprocess.Popen('florence',stdout=subprocess.PIPE, shell=True)
                subprocess.Popen('florence',stdout=subprocess.PIPE, shell=True)
            else:
                keyboard_button['bg']='grey85'
                try:
                    subprocess.Popen(['killall','florence'])
                except:
                    pass
                root.attributes('-fullscreen', True)

        idpos_click(0)

        cancel_button = Button(setidmc_labelframe, font=('Courier','10','bold'), bg="lavender", text="H???y" , height=3, width=14, borderwidth=0, command=cancel_click)
        cancel_button.place(x=316,y=320)
        save_button = Button(setidmc_labelframe, activebackground="gold", font=('Courier','10','bold'), bg="yellow", text="L??u", height=3, width=14, borderwidth=0, command=save_click)
        save_button.place(x=470,y=242)
        load_button = Button(setidmc_labelframe, font=('Courier','10','bold'), bg="lavender", text="Ch???nh s???a\nt???p", height=3, width=14, borderwidth=0, command=load_click)
        load_button.place(x=316,y=242)
        keyboard_button = Button(setidmc_labelframe, font=('Courier','10','bold'), bg="grey85", text="B??n ph??m", height=3, width=7, borderwidth=0, command=keyboard_click)
        keyboard_button.place(x=525,y=320)

        readerror_label = Label(mainscreen_labelframe, bg='red', text="H??? TH???NG L???I!", font=("Courier",14,"bold"))
        if(readerror==1):
            readerror_label.place(x=642,y=9)
            root.update_idletasks()
        else:
            readerror_label.place(x=1000,y=1000)
            root.update_idletasks()

    def power_click():
        try:
                subprocess.Popen(['killall','florence'])
        except:
            pass
        root.attributes('-fullscreen', True)

        home_canvas['bg'] = 'dodger blue'
        covid19_canvas['bg'] = 'dodger blue'
        viewresult_canvas['bg'] = 'dodger blue'
        setid_canvas['bg'] = 'dodger blue'
        config_canvas['bg'] = 'dodger blue'
        power_canvas['bg'] = 'white'

        powermc_labelframe = LabelFrame(mainscreen_labelframe, bg='white', width=624, height=478)
        powermc_labelframe.place(x=172,y=0)

        top_labelframe = LabelFrame(powermc_labelframe, bg='dodger blue', width=619, height=37)
        top_labelframe.place(x=1,y=1)

        viewresulttitle_label = Label(top_labelframe, bg='dodger blue', font=("Courier", 11, 'bold'), text="THO??T")
        viewresulttitle_label.place(x=280,y=7)

        probe_label = Label(powermc_labelframe, bg='white', fg = "grey80", font=("Courier", 10, 'bold'), text="Probe version")
        probe_label.place(x=5,y=449)

        logo_img = Image.open('/home/pi/Spotcheck/logo.png')
        logo_width, logo_height = logo_img.size
        scale_percent = 30
        width = int(logo_width * scale_percent / 100)
        height = int(logo_height * scale_percent / 100)
        display_img = logo_img.resize((width,height))
        image_select = ImageTk.PhotoImage(display_img)
        logo_label = Label(mainscreen_labelframe, bg='white',image=image_select)
        logo_label.image = image_select
        logo_label.place(x=650,y=442)

        power_labelframe = LabelFrame(powermc_labelframe, bg='white', width=405, height=120)
        power_labelframe.place(x=106,y=200)
        def shutdown_click():
            os.system("sudo shutdown -h now")
        def restart_click():
            os.system("sudo shutdown -r now")
        def exit_click():
            root.destroy()
        exit_button = Button(power_labelframe, fg='white', activebackground="dodger blue", font=('Courier','10','bold'), bg="blue4", text="????ng ???ng d???ng", height=5, width=12, borderwidth=0, command=exit_click)
        exit_button.place(x=9,y=12)
        shutdown_button = Button(power_labelframe, fg='white', activebackground="red", font=('Courier','10','bold'), bg="red3", text="T???t ngu???n", height=5, width=12, borderwidth=0, command=shutdown_click)
        shutdown_button.place(x=139,y=12)
        restart_button = Button(power_labelframe, fg='white', activebackground="lawn green", font=('Courier','10','bold'), bg="green", text="Kh???i ?????ng l???i", height=5, width=12, borderwidth=0, command=restart_click)
        restart_button.place(x=269,y=12)

        readerror_label = Label(mainscreen_labelframe, bg='red', text="H??? TH???NG L???I!", font=("Courier",14,"bold"))
        if(readerror==1):
            readerror_label.place(x=642,y=9)
            root.update_idletasks()
        else:
            readerror_label.place(x=1000,y=1000)
            root.update_idletasks()

    def config_click():
        try:
                subprocess.Popen(['killall','florence'])
        except:
            pass
        root.attributes('-fullscreen', True)

        home_canvas['bg'] = 'dodger blue'
        covid19_canvas['bg'] = 'dodger blue'
        viewresult_canvas['bg'] = 'dodger blue'
        setid_canvas['bg'] = 'dodger blue'
        config_canvas['bg'] = 'white'
        power_canvas['bg'] = 'dodger blue'

        global server_on, ftp_ip , ftp_user, ftp_password, ftp_folder

        configmc2_labelframe = LabelFrame(mainscreen_labelframe, bg='white', width=624, height=478)
        configmc2_labelframe.place(x=172,y=0)

        top_labelframe = LabelFrame(configmc2_labelframe, bg='dodger blue', width=619, height=37)
        top_labelframe.place(x=1,y=1)

        configtitle_label = Label(top_labelframe, bg='dodger blue', font=("Courier", 11, 'bold'), text="SERVER")
        configtitle_label.place(x=267,y=7)

        ip_label = Label(configmc2_labelframe, bg='white', text='?????a ch??? IP', font=('Courier',13,'bold'))
        ip_label.place(x=50,y=157)
        user_label = Label(configmc2_labelframe, bg='white', text='T??n ????ng nh???p', font=('Courier',13,'bold'))
        user_label.place(x=50,y=207)
        password_label = Label(configmc2_labelframe, bg='white', text='M???t kh???u', font=('Courier',13,'bold'))
        password_label.place(x=50,y=257)
        folder_label = Label(configmc2_labelframe, bg='white', text='???????ng d???n th?? m???c', font=('Courier',13,'bold'))
        folder_label.place(x=50,y=307)

        ip_entry = Entry(configmc2_labelframe,width=28, font=('Courier',14))
        ip_entry.place(x=253,y=155)
        user_entry = Entry(configmc2_labelframe,width=28, font=('Courier',14))
        user_entry.place(x=253,y=205)
        password_entry = Entry(configmc2_labelframe,width=28, show='???', font=('Courier',14))
        password_entry.place(x=253,y=255)
        folder_entry = Entry(configmc2_labelframe,width=28, font=('Courier',14))
        folder_entry.place(x=253,y=305)

        def on_click():
            on_button['bg']='lawn green'
            on_button['fg'] = 'black'
            off_button['bg']='grey88'
            off_button['fg'] = 'grey70'
            ip_entry['state'] = 'normal'
            user_entry['state'] = 'normal'
            password_entry['state'] = 'normal'
            folder_entry['state'] = 'normal'
            ip_entry.delete(0,END)
            ip_entry.insert(0,ftp_ip)
            user_entry.insert(0,ftp_user)
            folder_entry.insert(0,ftp_folder)

        def off_click():
            off_button['bg']='lawn green'
            off_button['fg'] = 'black'
            on_button['bg']='grey88'
            on_button['fg'] = 'grey70'
            ip_entry.delete(0,END)
            user_entry.delete(0,END)
            password_entry.delete(0,END)
            folder_entry.delete(0,END)
            ip_entry['state'] = 'disabled'
            user_entry['state'] = 'disabled'
            password_entry['state'] = 'disabled'
            folder_entry['state'] = 'disabled'


        if(server_on==1):
            on_button = Button(configmc2_labelframe, bg="lawn green", text="B???t", borderwidth=0, height=2, width=7,command=on_click)
            on_button.place(x=302,y=85)
            off_button = Button(configmc2_labelframe, bg="grey88",fg='grey70', text="T???t", borderwidth=0, height=2, width=7,command=off_click)
            off_button.place(x=220,y=85)
            ip_entry.insert(0,ftp_ip)
            user_entry.insert(0,ftp_user)
            folder_entry.insert(0,ftp_folder)

        else:
            on_button = Button(configmc2_labelframe, bg="grey88", fg='grey70', text="B???t", borderwidth=0, height=2, width=7, command=on_click)
            on_button.place(x=302,y=85)
            off_button = Button(configmc2_labelframe, bg="lawn green", text="T???t", borderwidth=0, height=2, width=7, command=off_click)
            off_button.place(x=220,y=85)
            ip_entry.delete(0,END)
            user_entry.delete(0,END)
            password_entry.delete(0,END)
            folder_entry.delete(0,END)
            ip_entry['state'] = 'disabled'
            user_entry['state'] = 'disabled'
            password_entry['state'] = 'disabled'
            folder_entry['state'] = 'disabled'

        def save_click():
            msg = messagebox.askquestion("L??u ", "B???n c?? mu???n l??u c??i ?????t ?")
            if(msg=='yes'):
                if(on_button['bg']=='lawn green'):
                    ip_set = ip_entry.get()
                    user_set = user_entry.get()
                    password_set = password_entry.get()
                    folder_set = folder_entry.get()
                    if(ip_set==''):
                        messagebox.showwarning("","B???n ch??a nh???p IP !")
                    elif(user_set==''):
                        messagebox.showwarning("","B???n ch??a nh???p T??n ????ng nh???p !")
                    elif(ip_set==''):
                        messagebox.showwarning("","B???n ch??a nh???p M???t kh???u !")
                    elif(folder_set==''):
                        messagebox.showwarning("","B???n ch??a nh???p ???????ng d???n th?? m???c !")
                    else:
                        try:
                            ftp = FTP(ip_set, user_set, password_set)
                            ftp.cwd(folder_set)
                            ftp.quit()
                            tc= open("/home/pi/Spotcheck/.server.txt","w")
                            tc.writelines('1\n')
                            tc.writelines(ip_set+"\n")
                            tc.writelines(user_set+"\n")
                            tc.writelines(password_set+"\n")
                            tc.writelines(folder_set+"\n")
                            global server_on, ftp_ip , ftp_user, ftp_password, ftp_folder
                            server_on = 1
                            ftp_ip = ip_set
                            ftp_user = user_set
                            ftp_password = password_set
                            ftp_folder = folder_set
                            messagebox.showinfo("", "???? l??u xong !")
                        except Exception as e :
                            error = messagebox.showwarning("Kh??ng th??? k???t n???i ?????n Server !",str(e))
                            if(error=='ok'):
                                pass
                else:
                    tc= open("/home/pi/Spotcheck/.server.txt","w")
                    tc.writelines('0\n')
                    tc.writelines("\n")
                    tc.writelines("\n")
                    tc.writelines("\n")
                    tc.writelines("\n")
                    server_on = 0
                    messagebox.showinfo("", "???? l??u xong !")

        def keyboard_click():
            if(keyboard_button['bg']=='grey85'):
                keyboard_button['bg']='lawn green'
                try:
                    subprocess.Popen(['killall','florence'])
                except:
                    pass
                root.attributes('-fullscreen', False)
                subprocess.Popen('florence',stdout=subprocess.PIPE, shell=True)
                subprocess.Popen('florence',stdout=subprocess.PIPE, shell=True)
            else:
                keyboard_button['bg']='grey85'
                try:
                    subprocess.Popen(['killall','florence'])
                except:
                    pass
                root.attributes('-fullscreen', True)

        keyboard_button = Button(configmc2_labelframe, font=('Courier','10','bold'), bg="grey85", text="B??n ph??m", height=3, width=7, borderwidth=0, command=keyboard_click)
        keyboard_button.place(x=530,y=374)
        save_button = Button(configmc2_labelframe, bg="yellow", text="L??u", borderwidth=0, height=3, width=10, command=save_click)
        save_button.place(x=253,y=374)

        readerror_label = Label(mainscreen_labelframe, bg='red', text="H??? TH???NG L???I!", font=("Courier",14,"bold"))
        if(readerror==1):
            readerror_label.place(x=642,y=9)
            root.update_idletasks()
        else:
            readerror_label.place(x=1000,y=1000)
            root.update_idletasks()

    def viewresult_click():
        try:
                subprocess.Popen(['killall','florence'])
        except:
            pass
        root.attributes('-fullscreen', True)

        home_canvas['bg'] = 'dodger blue'
        covid19_canvas['bg'] = 'dodger blue'
        viewresult_canvas['bg'] = 'white'
        setid_canvas['bg'] = 'dodger blue'
        config_canvas['bg'] = 'dodger blue'
        power_canvas['bg'] = 'dodger blue'

        viewresultmc_labelframe = LabelFrame(mainscreen_labelframe, bg='white', width=624, height=478)
        viewresultmc_labelframe.place(x=172,y=0)

        top_labelframe = LabelFrame(viewresultmc_labelframe, bg='dodger blue', width=619, height=37)
        top_labelframe.place(x=1,y=1)

        viewresulttitle_label = Label(top_labelframe, bg='dodger blue', font=("Courier", 11, 'bold'), text="HI???U CHU???N QUANG")
        viewresulttitle_label.place(x=225,y=7)

        step_label = Label(viewresultmc_labelframe,bg='white',fg="grey25", text="B?????c 1: Ki???m tra kh??ng m???u", font=("Courier",11, 'bold'))
        step_label.place(x=20,y=48)

        sampletable_labelframe = LabelFrame(viewresultmc_labelframe,bg='grey95', width=600, height=360)
        sampletable_labelframe.place(x=70,y=79)

        sample_label = list(range(48))
        h=0
        c=0
        for i in range(0,48):
            c+=1
            sample_label[i] = Label(sampletable_labelframe,bg="grey25",width=4,height=2)
            sample_label[i].grid(row=h,column=c,padx=3,pady=3)
            if(c>=6):
                h+=1
                c=0

        note_labelframe = LabelFrame(viewresultmc_labelframe,bg='grey95', width=220, height=272)
        note_labelframe.place(x=330,y=79)
        note1_label = Label(note_labelframe,bg="grey25",width=4,height=2)
        note1_label.place(x=2,y=2)
        note11_label = Label(note_labelframe,bg="grey95",font=("Courier",10),text="V??? tr?? kh??ng ?????t m???u")
        note11_label.place(x=50,y=12)
        note2_label = Label(note_labelframe,bg="lawn green",width=4,height=2)
        note2_label.place(x=2,y=47)
        note22_label = Label(note_labelframe,bg="grey95",font=("Courier",10),text="V??? tr?? ?????t m???u")
        note22_label.place(x=50,y=57)

        button_labelframe = LabelFrame(viewresultmc_labelframe,bg='grey95', width=220, height=78)
        button_labelframe.place(x=330,y=356 )

        def check_click():
            global thr_set
            check_button.place_forget()
            cprocess_label = Label(button_labelframe, bg="grey95", fg='blue', text='??ang x??? l??...', font=("Courier",11))
            cprocess_label.place(x=44,y=28)
            root.update_idletasks()
            send_data = 'P'
            ser.write(send_data.encode())

            if(ser.in_waiting>0):
                receive_data = ser.readline().decode('utf-8').rstrip()
                print("Data received:", receive_data)
                if(receive_data=='C'):
                    global wait
                    wait = 1

            while(wait!=1):
                if(ser.in_waiting>0):
                    receive_data = ser.readline().decode('utf-8').rstrip()
                    print("Data received:", receive_data)
                    if(receive_data=='C'):
                        wait = 1
                        break
            while(wait==1):
                if(step_label["text"] == "B?????c 1: Ki???m tra kh??ng m???u"):
                    try:
                        camera_capture("/home/pi/Spotcheck/Kiem tra do sang/a1.jpg")
                    except Exception as e :
                        error = messagebox.showerror(str(e), "L???i: Err 03", icon = "error")
                        if(error=='ok'):
                            root.destroy()

                    check_list=list(range(48))
                    check_list, check_img = process_image("/home/pi/Spotcheck/Kiem tra do sang/a1.jpg")

                    output = "/home/pi/Spotcheck/Kiem tra do sang/a2.jpg"
                    cv2.imwrite(output, check_img)

                    workbook = Workbook()
                    sheet = workbook.active
                    sheet["A2"] = "A"
                    sheet["A3"] = "B"
                    sheet["A4"] = "C"
                    sheet["A5"] = "D"
                    sheet["A6"] = "E"
                    sheet["A7"] = "F"
                    sheet["A8"] = "G"
                    sheet["A9"] = "H"
                    sheet["B1"] = "1"
                    sheet["C1"] = "2"
                    sheet["D1"] = "3"
                    sheet["E1"] = "4"
                    sheet["F1"] = "5"
                    sheet["G1"] = "6"
                    for i in range(0,48):
                        if(i<6):
                            pos = str(chr(65+i+1)) + "2"
                        if(i>=6 and i<12):
                            pos = str(chr(65+i-5)) + "3"
                        if(i>=12 and i<18):
                            pos = str(chr(65+i-11)) + "4"
                        if(i>=18 and i<24):
                            pos = str(chr(65+i-17)) + "5"
                        if(i>=24 and i<30):
                            pos = str(chr(65+i-23)) + "6"
                        if(i>=30 and i<36):
                            pos = str(chr(65+i-29)) + "7"
                        if(i>=36 and i<42):
                            pos = str(chr(65+i-35)) + "8"
                        if(i>=42):
                            pos = str(chr(65+i-41)) + "9"

                        sheet[pos] = check_list[i]

                    workbook.save("/home/pi/Spotcheck/Kiem tra do sang/gt1.xlsx")

                    check_average = round(sum(check_list)/len(check_list),1)
                    check_err=0
                    for i in range(0,48):
                        if(check_list[i] > tl[i]+tl[i]*15/100 or check_list[i] < tl[i]-tl[i]*15/100):
                            if(check_err!=1):
                                check_err = 1

                    global readerror
                    if(check_average > average_max or check_average < average_min):
                        cprocess_label['text'] = "L???i: ERR 01"
                        readerror = 1
                    elif(check_err==1):
                        readerror = 1
                        check_err=0
                        cprocess_label['text'] = "L???i: ERR 02"
                    else:
                        thr_set = round(check_average*a+b,1)
                        cprocess_label.place_forget()
                        check_button.place(x=53,y=14)
                        c=0
                        for i in range(0,48):
                            c+=1
                            if(c==2):
                                sample_label[i]['bg']='lawn green'
                                sample_label[i]['text']='N'
                            if(c==3):
                                sample_label[i]['bg']='lawn green'
                                sample_label[i]['text']='>30'
                            if(c==4):
                                sample_label[i]['bg']='lawn green'
                                sample_label[i]['text']='>25'
                            if(c==5):
                                sample_label[i]['bg']='lawn green'
                                sample_label[i]['text']='<25'
                            if(c>=6):
                                c=0
                        step_label["text"] = "B?????c 2: Ki???m tra FAM"

                else:
                    try:
                        camera_capture("/home/pi/Spotcheck/Kiem tra do sang/a3.jpg")
                    except Exception as e :
                        error = messagebox.showerror(str(e), "L???i: Err 03", icon = "error")
                        if(error=='ok'):
                            root.destroy()

                    check_list=list(range(48))
                    check_list, check_img = process_image("/home/pi/Spotcheck/Kiem tra do sang/a3.jpg")

                    output = "/home/pi/Spotcheck/Kiem tra do sang/a4.jpg"
                    cv2.imwrite(output, check_img)

                    workbook = Workbook()
                    sheet = workbook.active
                    sheet["A2"] = "A"
                    sheet["A3"] = "B"
                    sheet["A4"] = "C"
                    sheet["A5"] = "D"
                    sheet["A6"] = "E"
                    sheet["A7"] = "F"
                    sheet["A8"] = "G"
                    sheet["A9"] = "H"
                    sheet["B1"] = "1"
                    sheet["C1"] = "2"
                    sheet["D1"] = "3"
                    sheet["E1"] = "4"
                    sheet["F1"] = "5"
                    sheet["G1"] = "6"
                    for i in range(0,48):
                        if(i<6):
                            pos = str(chr(65+i+1)) + "2"
                        if(i>=6 and i<12):
                            pos = str(chr(65+i-5)) + "3"
                        if(i>=12 and i<18):
                            pos = str(chr(65+i-11)) + "4"
                        if(i>=18 and i<24):
                            pos = str(chr(65+i-17)) + "5"
                        if(i>=24 and i<30):
                            pos = str(chr(65+i-23)) + "6"
                        if(i>=30 and i<36):
                            pos = str(chr(65+i-29)) + "7"
                        if(i>=36 and i<42):
                            pos = str(chr(65+i-35)) + "8"
                        if(i>=42):
                            pos = str(chr(65+i-41)) + "9"

                        if(id_list[i]=='N/A'):
                            sheet[pos] = 'N/A'
                        else:
                            sheet[pos] = check_list[i]

                    workbook.save("/home/pi/Spotcheck/Kiem tra do sang/gt2.xlsx")

                    c=0
                    for i in range(0,48):
                        c+=1
                        sample_label[i]['text']=''
                        if(check_list[i]<float(thr_set)):
                            sample_label[i]['bg'] = 'green3'
                        if(check_list[i] < hs_ct1*float(thr_set)):
                            sample_label[i]['bg'] = 'yellow'
                        elif(check_list[i] >= hs_ct1*float(thr_set) and check_list[i] < hs_ct2*float(thr_set)):
                            sample_label[i]['bg'] = 'orange'
                        else:
                            sample_label[i]['bg'] = 'red'
                        if(c==1 or c==6):
                            sample_label[i]['bg'] = 'grey25'
                        if(c>=6):
                            c=0
                    c=0
                    err_green = 0
                    err_yellow = 0
                    err_orange = 0
                    err_red = 0
                    for i in range(0,48):
                        c+=1
                        if(c==2):
                            if(sample_label[i]['bg']=='orange'or sample_label[i]['bg']=='red'):
                                err_green+=3
                            elif(sample_label[i]['bg']=='yellow'):
                                err_green+=1
                        if(c==3):
                            if(sample_label[i]['bg']=='red'):
                                err_yellow+=3
                            elif(sample_label[i]['bg']=='orange' or sample_label[i]['bg']=='green3'):
                                err_yellow+=1
                        if(c==4):
                            if(sample_label[i]['bg']=='green3'):
                                err_orange+=3
                            elif(sample_label[i]['bg']=='yellow' or sample_label[i]['bg']=='red'):
                                err_orange+=1
                        if(c==5):
                            if(sample_label[i]['bg']=='green3' or sample_label[i]['bg']=='yellow'):
                                err_red+=3
                            elif(sample_label[i]['bg']=='orange'):
                                err_red+=1
                        if(c>=6):
                            c=0

                    print(err_green)
                    print(err_yellow)
                    print(err_orange)
                    print(err_red)

                    if(err_green>1 or err_yellow>2 or err_orange>2 or err_red>1):
                        readerror = 1
                        cprocess_label['text'] = "H??? th???ng l???i"
                        cprocess_label['fg'] = "red"
                        step_label['text'] = "Ho??n th??nh"
                    else:
                        readerror = 0
                        readerror_label = Label(mainscreen_labelframe, bg='dodger blue',fg='dodger blue', text="H??? TH???NG L???I!", font=("Courier",14,"bold"))
                        readerror_label.place(x=642,y=9)
                        root.update_idletasks()
                        cprocess_label['text'] = "H??? th???ng ???n ?????nh"
                        cprocess_label['fg'] = "green3"
                        cprocess_label.place(x=30,y=28)
                        step_label['text'] = "Ho??n th??nh"

                    note3_label = Label(note_labelframe,bg="green3",width=4,height=2)
                    note3_label.place(x=50,y=92)
                    note33_label = Label(note_labelframe,bg="grey95",font=("Courier",10),text="??m t??nh")
                    note33_label.place(x=98,y=102)
                    note4_label = Label(note_labelframe,bg="yellow",width=4,height=2)
                    note4_label.place(x=50,y=137)
                    note44_label = Label(note_labelframe,bg="grey95",font=("Courier",10),text="Ct > 30")
                    note44_label.place(x=98,y=147)
                    note5_label = Label(note_labelframe,bg="orange",width=4,height=2)
                    note5_label.place(x=50,y=182)
                    note55_label = Label(note_labelframe,bg="grey95",font=("Courier",10),text="Ct > 25")
                    note55_label.place(x=98,y=192)
                    note6_label = Label(note_labelframe,bg="red",width=4,height=2)
                    note6_label.place(x=50,y=227)
                    note66_label = Label(note_labelframe,bg="grey95",font=("Courier",10),text="Ct < 25")
                    note66_label.place(x=98,y=237)

                    root.update_idletasks()

                wait = 0
        check_button = Button(button_labelframe, bg="lawn green", text="Ki???m tra", borderwidth=0, height=2, width=10,command=check_click)
        check_button.place(x=53,y=14)

        readerror_label = Label(mainscreen_labelframe, bg='red', text="H??? TH???NG L???I!", font=("Courier",14,"bold"))
        if(readerror==1):
            readerror_label.place(x=642,y=9)
            root.update_idletasks()
        else:
            readerror_label.place(x=1000,y=1000)
            root.update_idletasks()

    home_button = Button(sidebar_labelframe, bg="dodger blue", activebackground="dodger blue", text="TRANG CH??? ", fg='white', font=buttonFont, borderwidth=0, height=4, width=20,command=home_click)
    home_button.place(x=1,y=1)
    home_canvas = Canvas(sidebar_labelframe, bg="dodger blue", bd=0, highlightthickness=0, height=72, width=13)
    home_canvas.place(x=2,y=3)
    setid_button = Button(sidebar_labelframe, bg="dodger blue", activebackground="dodger blue", text="T???P\nM???U X??T NGHI???M", fg='white', font=buttonFont, borderwidth=0, height=4, width=20, command=setid_click)
    setid_button.place(x=1,y=80),
    setid_canvas = Canvas(sidebar_labelframe, bg="dodger blue", bd=0, highlightthickness=0, height=72, width=13)
    setid_canvas.place(x=2,y=82)
    covid19_button = Button(sidebar_labelframe, bg="dodger blue", activebackground="dodger blue", text="PH??N T??CH", fg='white', font=buttonFont, borderwidth=0, height=4, width=20, command=covid19_click)
    covid19_button.place(x=1,y=159)
    covid19_canvas = Canvas(sidebar_labelframe, bg="dodger blue", bd=0, highlightthickness=0, height=72, width=13)
    covid19_canvas.place(x=2,y=161)
    viewresult_button = Button(sidebar_labelframe, bg="dodger blue", activebackground="dodger blue", text="HI???U CHU???N\nQUANG", fg='white', font=buttonFont, borderwidth=0, height=4, width=20, command=viewresult_click)
    viewresult_button.place(x=1,y=238)
    viewresult_canvas = Canvas(sidebar_labelframe, bg="dodger blue", bd=0, highlightthickness=0, height=72, width=13)
    viewresult_canvas.place(x=2,y=240)
    config_button = Button(sidebar_labelframe, bg="dodger blue", activebackground="dodger blue", text="SERVER", fg='white', font=buttonFont, borderwidth=0, height=4, width=20, command=config_click)
    config_button.place(x=1,y=317)
    config_canvas = Canvas(sidebar_labelframe, bg="dodger blue", bd=0, highlightthickness=0, height=72, width=13)
    config_canvas.place(x=2,y=319)
    power_button = Button(sidebar_labelframe, bg="dodger blue", activebackground="dodger blue", text="THO??T", fg='white', font=buttonFont, borderwidth=0, height=4, width=20, command=power_click)
    power_button.place(x=1,y=396)
    power_canvas = Canvas(sidebar_labelframe, bg="dodger blue", bd=0, highlightthickness=0, height=72, width=13)
    power_canvas.place(x=2,y=398)

    global covid19clicked
    if(covid19clicked==1):
        covid19_click()
    else:
        home_click()
############################################################# MAIN SCREEN - END ####################################################################

############################################################ VIEW RESULT - START ###################################################################
def result():
    result_labelframe = LabelFrame(root, bg='white', width=800, height=600)
    result_labelframe.place(x=0,y=0)

    result1_labelframe = LabelFrame(result_labelframe, bg='dodger blue', width=795, height=54)
    result1_labelframe.place(x=0,y=424)

    path_label = Label(result_labelframe, bg='red',text=rsfile, font=('Courier',9), width=112)
    path_label.place(x=4,y=4)

    result_img = Image.open(rsfile)
    result_width, result_height = result_img.size
    scale_percent = 81
    width = int(result_width * scale_percent / 100)
    height = int(result_height * scale_percent / 100)
    display_img = result_img.resize((width,height))
    image_select = ImageTk.PhotoImage(display_img)
    result_label = Label(result_labelframe, bg='white',image=image_select)
    result_label.image = image_select
    result_label.place(x=72,y=26)

    def back_click():
        result_labelframe.place_forget()
        mainscreen()
    def open_click():
        global rsfile
        p = rsfile
#         a=0
#         for i in range(len(p)):
#             if(p[i]=='/'):
#                 a=i
#         oldpath = p[:a]
        rsfile = filedialog.askopenfilename(initialdir='/home/pi/Spotcheck Ket Qua/Probe/', filetypes=[('jpg file','*.jpg')])
        if rsfile is not None:
            if(rsfile[len(rsfile)-3:]=='jpg'):
                result_img = Image.open(rsfile)
                result_width, result_height = result_img.size
                scale_percent = 81
                width = int(result_width * scale_percent / 100)
                height = int(result_height * scale_percent / 100)
                display_img = result_img.resize((width,height))
                image_select = ImageTk.PhotoImage(display_img)

                result_label = Label(result_labelframe, bg='white',image=image_select)
                result_label.image = image_select
                result_label.place(x=72,y=26)

                path_label['text']=rsfile
            else:
                 pass
    back_button = Button(result_labelframe, bg="lavender", text="Tr??? l???i" , height=2, width=8, borderwidth=0, command=back_click)
    back_button.place(x=290,y=428)
    open_button = Button(result_labelframe, bg="lavender", text="Xem ti???p" , height=2, width=8, borderwidth=0, command=open_click)
    open_button.place(x=415,y=428)
############################################################# VIEW RESULT - END ####################################################################

########################################################### SET ID SCREEN - START ##################################################################
def setid():
    setid1_labelframe = LabelFrame(root, bg='white', width=800, height=600)
    setid1_labelframe.place(x=0,y=0)

    setid2_labelframe = LabelFrame(setid1_labelframe, bg='white', width=470, height=160)
    setid2_labelframe.place(x=320,y=5)

    idpos_label = Label(setid2_labelframe, bg='dodger blue', font=("Courier",24,"bold"))
    idpos_label.place(x=1,y=1)

    setidtable_labelframe = LabelFrame(setid1_labelframe,bg='ghost white', width=600, height=307)
    setidtable_labelframe.place(x=10,y=5)

    def idpos_click(n):
        if(idpos_button[n]['bg'] != 'lawn green'):
            for k in range (0,48):
                if(idpos_button[k]['bg'] != 'lawn green' and idpos_button[k]['bg'] != 'grey99'):
                    idpos_button[k]['bg'] = 'lavender'
                else:
                    idpos_button[k]['bg'] = 'lawn green'
            idpos_button[n]['bg'] = 'dodger blue'
        else:
            for k in range (0,48):
                if(idpos_button[k]['bg'] != 'lawn green' and idpos_button[k]['bg'] != 'grey99'):
                    idpos_button[k]['bg'] = 'lavender'
                if(idpos_button[k]['bg'] == 'grey99'):
                    idpos_button[k]['bg'] = 'lawn green'
            idpos_button[n]['bg'] = 'grey99'

        def enter_entry(event):
            try:
                subprocess.Popen(['killall','florence'])
            except:
                pass
            root.attributes('-fullscreen', False)
            subprocess.Popen('florence',stdout=subprocess.PIPE, shell=True)
            subprocess.Popen('florence',stdout=subprocess.PIPE, shell=True)

        def ok_click(event=None):
            if(id_entry.get()==''):
                idpos_button[n]['bg'] = 'lavender'
                idpos_button[n]['text'] = '#'+str(n+1)
                msgbox = messagebox.showwarning(" ","B???n ch??a nh???p ID !")
            else:
                idpos_button[n]['text'] = id_entry.get()
                idpos_button[n]['bg'] = 'lawn green'
                try:
                    if(n==47):
                        idpos_click(0)
                    else:
                        idpos_click(n+1)
                except:
                    idpos_click(0)

        id_entry = Entry(setid2_labelframe,width=25, font=('Courier',14))
        if(idpos_button[n]['bg'] == 'grey99'):
            id_entry.insert(0,idpos_button[n]['text'])
        #id_entry.bind("<Button-1>", enter_entry)
        id_entry.bind("<Return>", ok_click)
        id_entry.place(x=50,y=70)
        id_entry.focus_set()

        setid_label = Label(setid2_labelframe, text='Nh???p m???u x??t nghi???m', bg='white', font=("Courier",15,"bold"))
        setid_label.place(x=48,y=43)

        if(n<8):
            idpos_label['text'] = str(chr(65+n)) + '1'
        if(n>=8 and n<16):
            idpos_label['text'] = str(chr(65+n-8)) + '2'
        if(n>=16 and n<24):
            idpos_label['text'] = str(chr(65+n-16)) + '3'
        if(n>=24 and n<32):
            idpos_label['text'] = str(chr(65+n-24)) + '4'
        if(n>=32 and n<40):
            idpos_label['text'] = str(chr(65+n-32)) + '5'
        if(n>=40):
            idpos_label['text'] = str(chr(65+n-40)) + '6'

        ok_button = Button(setid2_labelframe, font=('Courier','12','bold'), bg="lavender", text="X??c nh???n", height=2, width=8, borderwidth=0, command=ok_click)
        ok_button.place(x=340,y=58)

    idpos_button = list(range(48))
    h=-1
    c=0
    for i in range(0,48):
        h+=1
        if(i%8==0 and i!=0):
            h=0
            c+=1
        idpos_button[i] = Button(setidtable_labelframe, bg='lavender', activebackground="white", justify='left', borderwidth=0, text='#'+str(i+1), width=2, height=2)
        idpos_button[i]['command'] = partial(idpos_click,i)
        idpos_button[i].grid(row=h,column=c,padx=4,pady=4)
        # if(i==46):
        #     idpos_button[i]['state']='disabled'
        #     idpos_button[i]['bg']= 'green'
        #     idpos_button[i]['text']= 'N'
        # if(i==47):
        #     idpos_button[i]['state']='disabled'
        #     idpos_button[i]['bg']= 'red'
        #     idpos_button[i]['text']= 'P'

    def cancel_click():
        msg = messagebox.askquestion("H???y", "B???n mu???n h???y m?? kh??ng l??u l???i t???p ?")
        if(msg=="yes"):
            setid1_labelframe.place_forget()
            mainscreen()

    def save_click():
        workbook = load_workbook("/home/pi/Spotcheck/template.xlsm", keep_vba = True)
        sheet = workbook.active
        # for i in range(0,48):
        #     #pos = "C"+str(i+3)
        #     if(i<8):
        #         pos = 'B'+ str(i+2)
        #     if(i>=8 and i<16):
        #         pos = 'C'+ str(i-6)
        #     if(i>=16 and i<24):
        #         pos = 'D'+ str(i-14)
        #     if(i>=24 and i<32):
        #         pos = 'E'+ str(i-22)
        #     if(i>=32 and i<40):
        #         pos = 'F'+ str(i-30)
        #     if(i>=40):
        #         pos = 'G'+str(i-38)

        for i in range(0,48):
            pos = "B" + str(i+12)
            if(idpos_button[i]['bg']=='lawn green' or idpos_button[i]['bg']=='grey99'):
                sheet[pos] = idpos_button[i]['text']
            else:
                sheet[pos] = 'N/A'

        # sheet['B58']='NEGC'
        # sheet['B59']='POSC'

        msg = messagebox.askquestion("L??u ", "B???n c?? mu???n l??u t???p ?")
        if(msg=='yes'):
            f = filedialog.asksaveasfilename(initialdir='/home/pi/Desktop/Spotcheck ID/',defaultextension='.xlsx')
            if f is not None:
                d=0
                for i in range(len(f)):
                    if(f[i]=='/'):
                        d=i+1
                filename = f[d:(len(f)-5)]
                print(filename)
                if(len(filename)<=30):
                    workbook.save(f)
                    try:
                        subprocess.Popen(['killall','florence'])
                    except:
                        pass
                    root.attributes('-fullscreen', True)

                    msg = messagebox.askquestion(' ','???? l??u xong!\nB???n c?? mu???n t???o t???p m???i ?')
                    if(msg=='yes'):
                        setid()
                    else:
                        setid1_labelframe.place_forget()
                        mainscreen()

                else:
                    messagebox.showerror("L???i", "T??n t???p kh??ng v?????t qu?? 30 k?? t??? !")

    def load_click():
        idfile = filedialog.askopenfilename(initialdir='/home/pi/Desktop/Spotcheck ID', filetypes=[('Excel file','*.xlsm *.xlsx *.xls')])
        if idfile is not None:
            if(idfile[len(idfile)-4:]=='xlsx' or idfile[len(idfile)-4:]=='xlsm' or idfile[len(idfile)-3:]=='xls'):
                workbook = openpyxl.load_workbook(idfile)
                sheet = workbook.active
                idfile_list = list(range(48))

                # for i in range(0,48):
                #     if(i<8):
                #         pos = 'B'+ str(i+2)
                #     if(i>=8 and i<16):
                #         pos = 'C'+ str(i-6)
                #     if(i>=16 and i<24):
                #         pos = 'D'+ str(i-14)
                #     if(i>=24 and i<32):
                #         pos = 'E'+ str(i-22)
                #     if(i>=32 and i<40):
                #         pos = 'F'+ str(i-30)
                #     if(i>=40):
                #         pos = 'G'+str(i-38)

                for i in range(0,48):
                    pos = 'B' + str(i+12)
                    idfile_list[i] = sheet[pos].value
                    idpos_button[i]['text'] = idfile_list[i]
                    if(idpos_button[i]['text']!='N/A'):
                        idpos_button[i]['bg']='lawn green'
                    # if(i==46):
                    #     idpos_button[i]['bg']= 'green'
                    # if(i==47):
                    #     idpos_button[i]['bg']= 'red'
        else:
            pass
    def keyboard_click():
        if(keyboard_button['bg']=='grey85'):
            keyboard_button['bg']='lawn green'
            try:
                subprocess.Popen(['killall','florence'])
            except:
                pass
            root.attributes('-fullscreen', False)
            subprocess.Popen('florence',stdout=subprocess.PIPE, shell=True)
            subprocess.Popen('florence',stdout=subprocess.PIPE, shell=True)
        else:
            keyboard_button['bg']='grey85'
            try:
                subprocess.Popen(['killall','florence'])
            except:
                pass
            root.attributes('-fullscreen', True)

    idpos_click(0)

    cancel_button = Button(setid1_labelframe, font=('Courier','12','bold'), bg="lavender", text="H???y" , height=3, width=11, borderwidth=0, command=cancel_click)
    cancel_button.place(x=653,y=170)
    save_button = Button(setid1_labelframe, activebackground="gold", font=('Courier','12','bold'), bg="yellow", text="L??u", height=3, width=11, borderwidth=0, command=save_click)
    save_button.place(x=487,y=170)
    load_button = Button(setid1_labelframe, font=('Courier','12','bold'), bg="lavender", text="Ch???nh s???a\nt???p s???n c??", height=3, width=11, borderwidth=0, command=load_click)
    load_button.place(x=320,y=170)
    keyboard_button = Button(setid1_labelframe, font=('Courier','10','bold'), bg="grey85", text="B??n ph??m", height=3, width=7, borderwidth=0, command=keyboard_click)
    keyboard_button.place(x=706,y=374)
############################################################ SET ID SCREEN - END ###################################################################

###################################################### SET TEMPERATURES SCREEN - START #############################################################
def settemp():
    if(covid19clicked==1):
        fr = open("/home/pi/Spotcheck/covid19saved.txt","r")
    if(tbclicked==1):
        fr = open("/home/pi/Spotcheck/tbsaved.txt","r")
    if(spotcheckclicked==1):
        fr = open("/home/pi/Spotcheck/scsaved.txt","r")
    if(shrimpclicked==1):
        fr = open("/home/pi/Spotcheck/shrimpsaved.txt","r")
    t1 = fr.readline()[3:5]
    t2 = fr.readline()[3:5]
    t3 = fr.readline()[3:5]
#     thr1 = fr.readline()[5:9]
#     thr2 = fr.readline()[5:9]
#     thr3l = fr.readline()[6:10]
#     thr3h = fr.readline()[6:10]
    global samples
    samples=0
    settemp_labelframe = LabelFrame(root, bg='white', width=800, height=600)
    settemp_labelframe.place(x=0,y=0)
    settemptop_labelframe = LabelFrame(settemp_labelframe, bg='white', width=798, height=350)
    settemptop_labelframe.place(x=0,y=52)
    keypad_labelframe = LabelFrame(settemptop_labelframe, bg='white', width=285, height=323)
    keypad_labelframe.place(x=501,y=11)
    title_labelframe = LabelFrame(settemp_labelframe, bg='dodger blue', width=798, height=50)
    title_labelframe.place(x=0,y=0)
    settemp_label = Label(settemp_labelframe, bg='dodger blue', fg='black', text='C??I ?????T NHI???T ?????', font=("Courier",17,'bold'), width=20, height=1 )
    settemp_label.place(x=260,y=12)

    def numpad_click(btn):
        text = "%s" % btn
        if (text!="X??a" and text!='M???c ?????nh'):
            if(entry_num==1):
                t1_entry.insert(END, text)
            if(entry_num==2):
                t2_entry.insert(END, text)
            if(entry_num==3):
                t3_entry.insert(END, text)
#             if(entry_num==4):
#                 thr1_entry.insert(END, text)
#             if(entry_num==5):
#                 thr2_entry.insert(END, text)
#             if(entry_num==6):
#                 thr3l_entry.insert(END, text)
#             if(entry_num==7):
#                 thr3h_entry.insert(END, text)
        if text == 'X??a':
            if(entry_num==1):
                t1_entry.delete(0, END)
            if(entry_num==2):
                t2_entry.delete(0, END)
            if(entry_num==3):
                t3_entry.delete(0, END)
#             if(entry_num==4):
#                 thr1_entry.delete(0, END)
#             if(entry_num==5):
#                 thr2_entry.delete(0, END)
#             if(entry_num==6):
#                 thr3l_entry.delete(0, END)
#             if(entry_num==7):
#                 thr3h_entry.delete(0, END)
        if text == 'M???c ?????nh':
            if(entry_num==1):
                t1_entry.delete(0, END)
                t1_entry.insert(END, t1)
            if(entry_num==2):
                t2_entry.delete(0, END)
                t2_entry.insert(END, t2)
            if(entry_num==3):
                t3_entry.delete(0, END)
                t3_entry.insert(END, t3)
#             if(entry_num==4):
#                 thr1_entry.delete(0, END)
#                 thr1_entry.insert(END, 25)
#             if(entry_num==5):
#                 thr2_entry.delete(0, END)
#                 thr2_entry.insert(END, 25)
#             if(entry_num==6):
#                 thr3l_entry.delete(0, END)
#                 thr3l_entry.insert(END, 25)
#             if(entry_num==7):
#                 thr3h_entry.delete(0, END)
#                 thr3h_entry.insert(END, 25)

    def numpad():
        global numpad_labelframe
        numpad_labelframe = LabelFrame(keypad_labelframe, bg="white", width=385, height=395)
        numpad_labelframe.place(x=2,y=1)
        button_list = ['7',     '8',      '9',
                       '4',     '5',      '6',
                       '1',     '2',      '3',
                       '0',     'X??a', 'M???c ?????nh']
        r = 1
        c = 0
        n = 0
        btn = list(range(len(button_list)))
        for label in button_list:
            cmd = partial(numpad_click, label)
            btn[n] = Button(numpad_labelframe, text=label, font=font.Font(family='Helvetica', size=10, weight='bold'), width=9, height=4, command=cmd)
            btn[n].grid(row=r, column=c, padx=0, pady=0)
            n += 1
            c += 1
            if (c == 3):
                c = 0
                r += 1

    temp_labelframe = LabelFrame(settemptop_labelframe, text='NHI???T ?????', bg='white', width=490, height=180)
    temp_labelframe.place(x=3,y=2)
#     thres_labelframe = LabelFrame(settemptop_labelframe, text='THRESHOLD', bg='white', width=490, height=149)
#     thres_labelframe.place(x=3,y=185)

    cir_img = Image.open('/home/pi/Spotcheck/cir.png')
    cir_width, cir_height = cir_img.size
    scale_percent = 14
    width = int(cir_width * scale_percent / 100)
    height = int(cir_height * scale_percent / 100)
    display_img = cir_img.resize((width,height))
    image_select = ImageTk.PhotoImage(display_img)
    t1cir_label = Label(temp_labelframe, bg='white', image=image_select)
    t1cir_label.image = image_select
    t1cir_label.place(x=5,y=5)
    t2cir_label = Label(temp_labelframe, bg='white', image=image_select)
    t2cir_label.image = image_select
    t2cir_label.place(x=170,y=5)
    t3cir_label = Label(temp_labelframe, bg='white', image=image_select)
    t3cir_label.image = image_select
    t3cir_label.place(x=335,y=5)
#     graycir_img = Image.open('graycir.png')
#     graycir_width, cir_height = cir_img.size
#     scale_percent = 16
#     width = int(cir_width * scale_percent / 100)
#     height = int(cir_height * scale_percent / 100)
#     display_img = graycir_img.resize((width,height))
#     image_select = ImageTk.PhotoImage(display_img)
#     t4cir_label = Label(settemptop_labelframe, bg='white', image=image_select)
#     t4cir_label.image = image_select
#     t4cir_label.place(x=275,y=175)

    def entryt1_click(event):
        global numpad_labelframe
        global entry_num
        entry_num = 1
        numpad()
    def entryt2_click(event):
        global numpad_labelframe
        global entry_num
        entry_num = 2
        numpad()
    def entryt3_click(event):
        global numpad_labelframe
        global entry_num
        entry_num = 3
        numpad()
#     def entrythr1_click(event):
#         global numpad_labelframe
#         global entry_num
#         entry_num = 4
#         numpad()
#     def entrythr2_click(event):
#         global numpad_labelframe
#         global entry_num
#         entry_num = 5
#         numpad()
#     def entrythr3l_click(event):
#         global numpad_labelframe
#         global entry_num
#         entry_num = 6
#         numpad()
#     def entrythr3h_click(event):
#         global numpad_labelframe
#         global entry_num
#         entry_num = 7
#         numpad()

    t1_label = Label(temp_labelframe, bg='white', text='T1', fg='black', font=("Courier",20,"bold"))
    t1_label.place(x=15, y=14)
    t1oc_label = Label(temp_labelframe, bg='white', text=chr(176)+'C', fg='red', font=("Courier", 11,"bold"))
    t1oc_label.place(x=107, y=55)
    t1_entry = Entry(temp_labelframe, width=2, justify='center', bg='white', borderwidth=0, fg ='grey32', font=("Courier",36,"bold"))
    t1_entry.place(x=47,y=50)
    t1_entry.bind('<Button-1>', entryt1_click)
    t1_entry.insert(0,t1)

    t2_label = Label(temp_labelframe, bg='white', text='T2', fg='black', font=("Courier",20,"bold"))
    t2_label.place(x=180, y=14)
    t2oc_label = Label(temp_labelframe, bg='white', text=chr(176)+'C', fg='red', font=("Courier", 11,"bold"))
    t2oc_label.place(x=272, y=55)
    t2_entry = Entry(temp_labelframe, width=2, justify='center', bg='white', borderwidth=0, fg ='grey32', font=("Courier",36,"bold"))
    t2_entry.place(x=212,y=50)
    t2_entry.bind('<Button-1>', entryt2_click)
    t2_entry.insert(0,t2)

    t3_label = Label(temp_labelframe, bg='white', text='T3', fg='black', font=("Courier",20,"bold"))
    t3_label.place(x=345, y=14)
    t3oc_label = Label(temp_labelframe, bg='white', text=chr(176)+'C', fg='red', font=("Courier", 11,"bold"))
    t3oc_label.place(x=437, y=55)
    t3_entry = Entry(temp_labelframe, width=2, justify='center', bg='white', borderwidth=0, fg ='grey32', font=("Courier",36,"bold"))
    t3_entry.place(x=377,y=50)
    t3_entry.bind('<Button-1>', entryt3_click)
    t3_entry.insert(0,t3)

#     thr1_label = Label(thres_labelframe, bg='white', text='T1: ', fg='black', font=("Courier",24,"bold"))
#     thr1_label.place(x=60, y=7)
#     thr1_entry = Entry(thres_labelframe, width=4, justify='center', bg='white', borderwidth=0, fg ='grey32', font=("Courier",25,"bold"))
#     thr1_entry.place(x=118,y=7)
#     thr1_entry.bind('<Button-1>', entrythr1_click)
#     thr1_entry.insert(0,thr1)
#
#     thr2_label = Label(thres_labelframe, bg='white', text='T2: ', fg='black', font=("Courier",24,"bold"))
#     thr2_label.place(x=60, y=71)
#     thr2_entry = Entry(thres_labelframe, width=4, justify='center', bg='white', borderwidth=0, fg ='grey32', font=("Courier",25,"bold"))
#     thr2_entry.place(x=118,y=71)
#     thr2_entry.bind('<Button-1>', entrythr2_click)
#     thr2_entry.insert(0,thr2)
#
#     thr3l_label = Label(thres_labelframe, bg='white', text='T3-L: ', fg='black', font=("Courier",24,"bold"))
#     thr3l_label.place(x=235, y=7)
#     thr3l_entry = Entry(thres_labelframe, width=4, justify='center', bg='white', borderwidth=0, fg ='grey32', font=("Courier",25,"bold"))
#     thr3l_entry.place(x=330,y=7)
#     thr3l_entry.bind('<Button-1>', entrythr3l_click)
#     thr3l_entry.insert(0,thr3l)
#
#     thr3h_label = Label(thres_labelframe, bg='white', text='T3-H: ', fg='black', font=("Courier",24,"bold"))
#     thr3h_label.place(x=235, y=71)
#     thr3h_entry = Entry(thres_labelframe, width=4, justify='center', bg='white', borderwidth=0, fg ='grey32', font=("Courier",25,"bold"))
#     thr3h_entry.place(x=330,y=71)
#     thr3h_entry.bind('<Button-1>', entrythr3h_click)
#     thr3h_entry.insert(0,thr3h)

#     t4_label = Label(settemptop_labelframe, bg = 'white', text='T4', fg='grey67', font=("Courier",20,"bold"))
#     t4_label.place(x=286, y=185)

    def back_click():
        settemp_labelframe.place_forget()
        mainscreen()
    def thread():
        th1 = Thread(target = next_click)
        th1.start()
    def next_click():
        try:
            camera.close()
        except:
            pass
        settemp_labelframe.place_forget()
        global t1_set, t2_set, t3_set
#         global thr1_set, thr2_set, thr3l_set, thr3h_set
        t1_set = t1_entry.get()[0:2]
        t2_set = t2_entry.get()[0:2]
        t3_set = t3_entry.get()[0:2]
#         thr1_set = thr1_entry.get()[0:4]
#         thr2_set = thr2_entry.get()[0:4]
#         thr3l_set = thr3l_entry.get()[0:4]
#         thr3h_set = thr3h_entry.get()[0:4]

        global path5
        if os.path.exists(path5+"/nhiet-do.txt"):
            fc= open(path5+"/nhietdo.txt","w")
            fc.truncate(0)
            fc.writelines("T1="+t1_entry.get()[0:2]+"\n")
            fc.writelines("T2="+t2_entry.get()[0:2]+"\n")
            fc.writelines("T3="+t3_entry.get()[0:2]+"\n")
        else:
            fc= open(path5+"/nhietdo.txt","w+")
            fc.writelines("T1="+t1_entry.get()[0:2]+"\n")
            fc.writelines("T2="+t2_entry.get()[0:2]+"\n")
            fc.writelines("T3="+t3_entry.get()[0:2]+"\n")
        scanposition()
    def save_click():
        msg = messagebox.askquestion("L??u ch????ng tr??nh nhi???t", "B???n c?? mu???n l??u nhi???t ????? ?")
        if(msg=='yes'):
            messagebox.showinfo("","???? l??u xong !")
            if(covid19clicked==1):
                fw = open("/home/pi/Spotcheck/covid19saved.txt","w")
            if(tbclicked==1):
                fw = open("/home/pi/Spotcheck/tbsaved.txt","w")
            if(spotcheckclicked==1):
                fw = open("/home/pi/Spotcheck/scsaved.txt","w")
            if(shrimpclicked==1):
                fw = open("/home/pi/Spotcheck/shrimpsaved.txt","w")
            fw.truncate(0)
            fw.writelines("T1="+t1_entry.get()[0:2]+"\n")
            fw.writelines("T2="+t2_entry.get()[0:2]+"\n")
            fw.writelines("T3="+t3_entry.get()[0:2]+"\n")
#         if(len(thr1_entry.get())<=2):
#             fw.writelines("THR1="+thr1_entry.get()[0:2]+".0"+"\n")
#         else:
#             fw.writelines("THR1="+thr1_entry.get()[0:4]+"\n")
#         if(len(thr2_entry.get())<=2):
#             fw.writelines("THR1="+thr2_entry.get()[0:2]+".0"+"\n")
#         else:
#             fw.writelines("THR2="+thr2_entry.get()[0:4]+"\n")
#         if(len(thr3l_entry.get())<=2):
#             fw.writelines("THR1="+thr3l_entry.get()[0:2]+".0"+"\n")
#         else:
#             fw.writelines("THR3L="+thr3l_entry.get()[0:4]+"\n")
#         if(len(thr3h_entry.get())<=2):
#             fw.writelines("THR1="+thr3h_entry.get()[0:2]+".0"+"\n")
#         else:
#             fw.writelines("THR3H="+thr3h_entry.get()[0:4]+"\n")

    back_button = Button(settemp_labelframe, font=('Courier','12','bold'), bg="lavender", text="Tr??? l???i" , height=3, width=11, borderwidth=0, command=back_click)
    back_button.place(x=14,y=406)
    next_button = Button(settemp_labelframe, font=('Courier','12','bold'), bg="lavender", text="Ti???p theo", height=3, width=11, borderwidth=0, command=thread)
    next_button.place(x=647,y=406)
    save_button = Button(settemp_labelframe, activebackground="gold", font=('Courier','12','bold'), bg="yellow", text="L??u", height=3, width=11, borderwidth=0,command=save_click)
    save_button.place(x=332,y=406)
####################################################### SET TEMPERATURES SCREEN - END ##############################################################

######################################################### SAMPLES POSITION - START #################################################################
def scanposition():
    global path0
    global path1
    global path2
    global path3
    global path4
    global path5

    global ser
    ser.flushInput()
    ser.flushOutput()
    global scanpostion_labelframe
    scanposition_labelframe = LabelFrame(root, bg='white', width=800, height=480)
    scanposition_labelframe.place(x=0,y=0)
    scanposition1_labelframe = LabelFrame(scanposition_labelframe, bg='white', width=797, height=368)
    scanposition1_labelframe.place(x=0,y=41)
    scanposition2_labelframe = LabelFrame(scanposition_labelframe, bg='white', width=797, height=67)
    scanposition2_labelframe.place(x=0,y=409)
    title_labelframe = LabelFrame(scanposition_labelframe, bg='dodger blue', width=796, height=40)
    title_labelframe.place(x=0,y=0)
    scanposition_label = Label(scanposition_labelframe, bg='dodger blue', text='KI???M TRA H??? TH???NG', font=("Courier",12,'bold'), width=20, height=1 )
    scanposition_label.place(x=295,y=10)

    scan_img = Image.open('/home/pi/Spotcheck/scan.png')
    scan_width, scan_height = scan_img.size
    scale_percent = 70
    width = int(scan_width * scale_percent / 100)
    height = int(scan_height * scale_percent / 100)
    display_img = scan_img.resize((width,height))
    image_select = ImageTk.PhotoImage(display_img)
    scan_label = Label(scanposition1_labelframe, bg='white',image=image_select)
    scan_label.image = image_select
    scan_label.place(x=218,y=3)

    s = ttk.Style()
    s.theme_use('clam')
    s.configure("green.Horizontal.TProgressbar", foreground='green', background='green')
    scanposition_progressbar = ttk.Progressbar(root, orient = HORIZONTAL, style="green.Horizontal.TProgressbar", length = 200, mode = 'determinate')
    scanposition_progressbar.place(x=299,y=422)
    root.update_idletasks()

    def back_click():
        try:
            camera.close()
        except Exception:
            pass
        global wait
        wait = 0
        mainscreen()

    back_button = Button(scanposition2_labelframe, font=("Courier",10,'bold'), bg="lavender", text="Tr??? l???i" , height=3, width=11, borderwidth=0, command=back_click)
    back_button.place(x=1,y=1)
    process_label = Label(scanposition_labelframe, text='??ang x??? l?? ...', bg='white', font=("Courier",13))
    process_label.place(x=330,y=448)

    send_data = 'P'
    ser.write(send_data.encode())

    if(ser.in_waiting>0):
        receive_data = ser.readline().decode('utf-8').rstrip()
        print("Data received:", receive_data)
        scanposition_progressbar['value'] = 5
        root.update_idletasks()
        if(receive_data=='C'):
            global wait
            wait = 1
            scanposition_progressbar['value'] = 20
            root.update_idletasks()

    while(wait!=1):
        scanposition_progressbar['value'] = 2
        root.update_idletasks()
        if(ser.in_waiting>0):
            receive_data = ser.readline().decode('utf-8').rstrip()
            print("Data received:", receive_data)
            scanposition_progressbar['value'] = 10
            root.update_idletasks()
            if(receive_data=='C'):
                scanposition_progressbar['value'] = 20
                root.update_idletasks()
                wait = 1
                break
    while(wait==1):
        try:
            camera_capture(path4 + "/mau.jpg")
        except Exception as e :
            error = messagebox.showerror(str(e), "L???i: Err 03", icon = "error")
            if(error=='ok'):
                root.destroy()

        image = cv2.imread(path4 + "/mau.jpg")
        blur_img = cv2.fastNlMeansDenoisingColored(image.copy(),None,15,15,7,21)
        gray_img = cv2.cvtColor(blur_img, cv2.COLOR_BGR2GRAY)
        thresh, binary_img = cv2.threshold(gray_img.copy(), 40, maxval=255, type=cv2.THRESH_BINARY)
        contours, hierarchy = cv2.findContours(binary_img, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
        print("Number of contours: " + str(len(contours)))

        contours.sort(key=lambda data:sorting_xy(data))

        contour_img = np.zeros_like(gray_img)
        bourect0 = cv2.boundingRect(contours[0])
        bourect47 = cv2.boundingRect(contours[len(contours)-1])
        global start_point
        start_point = (bourect0[0]-9, bourect0[1]-9)
        global end_point
        end_point = (bourect47[0]+bourect47[2]+9, bourect47[1]+bourect47[3]+9)
        print('Start point:', start_point)
        print('End point:', end_point)
        fw= open("/home/pi/Spotcheck/coordinates2.txt",'w')
        fw.writelines("Start Point: " + str(start_point) + "\n")
        fw.writelines("End Point: " + str(end_point))

        scanposition_progressbar['value'] = 35
        root.update_idletasks()

        global pos_result
        pos_result, pos_image = process_image(path4 + "/mau.jpg")
        #pos_result, pos_image = process_image("/home/pi/Desktop/mau.jpg")
        scanposition_progressbar['value'] = 60
        root.update_idletasks()
        sleep(1)

        output = path4 + "/xu-ly-mau.jpg"
        cv2.imwrite(output, pos_image)
        scanposition_progressbar['value'] = 90
        root.update_idletasks()
        sleep(1)

        workbook0 = Workbook()
        sheet0 = workbook0.active

        sheet0["A2"] = "A"
        sheet0["A3"] = "B"
        sheet0["A4"] = "C"
        sheet0["A5"] = "D"
        sheet0["A6"] = "E"
        sheet0["A7"] = "F"
        sheet0["A8"] = "G"
        sheet0["A9"] = "H"
        sheet0["B1"] = "1"
        sheet0["C1"] = "2"
        sheet0["D1"] = "3"
        sheet0["E1"] = "4"
        sheet0["F1"] = "5"
        sheet0["G1"] = "6"
        for i in range(0,48):
            if(i<6):
                pos = str(chr(65+i+1)) + "2"
            if(i>=6 and i<12):
                pos = str(chr(65+i-5)) + "3"
            if(i>=12 and i<18):
                pos = str(chr(65+i-11)) + "4"
            if(i>=18 and i<24):
                pos = str(chr(65+i-17)) + "5"
            if(i>=24 and i<30):
                pos = str(chr(65+i-23)) + "6"
            if(i>=30 and i<36):
                pos = str(chr(65+i-29)) + "7"
            if(i>=36 and i<42):
                pos = str(chr(65+i-35)) + "8"
            if(i>=42):
                pos = str(chr(65+i-41)) + "9"

            sheet0[pos] = pos_result[i]

        average_value = round(sum(pos_result)/len(pos_result),1)
        sheet0['I2'] = "Average: " + str(average_value)
        sheet0['I3'] = "Threshold: " + str(round(average_value*a+b,1))
        workbook0.save(path4 + "/gia-tri.xlsx")

        scanresult_labelframe = LabelFrame(scanposition1_labelframe, bg='ghost white', width=528,height = 307)
        scanresult_labelframe.place(x=244,y=4)

        average_tl = round(sum(tl)/len(tl),1)
        hs_tmp = round(average_tl/average_value,2)

        label = list(range(48))
        def result_table(range_a, range_b, row_value):
            j=-1
            for i in range(range_a, range_b):
                j+=1
                if(i<6):
                    t='A'+ str(i+1)
                if(i>=6 and i<12):
                    t='B'+ str(i-5)
                if(i>=12 and i<18):
                    t='C'+ str(i-11)
                if(i>=18 and i<24):
                    t='D'+ str(i-17)
                if(i>=24 and i<30):
                    t='E'+ str(i-23)
                if(i>=30 and i<36):
                    t='F'+ str(i-29)
                if(i>=36 and i<42):
                    t='G'+ str(i-35)
                if(i>=42):
                    t='H'+ str(i-41)

                label[i] = Label(scanresult_labelframe, bg='white', text=t, width=5, height=2)
                label[i].grid(row=row_value,column=j,padx=3,pady=3)

        scanposition_progressbar['value'] = 100
        root.update_idletasks()

        result_table(0,6,0)
        result_table(6,12,1)
        result_table(12,18,2)
        result_table(18,24,3)
        result_table(24,30,4)
        result_table(30,36,5)
        result_table(36,42,6)
        result_table(42,48,7)

        global errors
        print("average_value: ",average_value)
        print("average_tl: ",average_tl)
        print("hs_tmp: ",hs_tmp)
        for i in range(0,48):
            if(pos_result[i] > tl[i]+tl[i]*15/100 or pos_result[i] < tl[i]-tl[i]*15/100):
                label[i]['bg']="grey50"
                if(errors != 2):
                    errors=2
            elif(pos_result[i] >= tl[i]*2):
                label[i]['bg']="grey50"
                if(errors != 200):
                    errors=200
            else:
                label[i]['bg']="dodger blue"

        if(average_value > average_max or average_value < average_min):
            for i in range(0,48):
                label[i]['bg']='grey50'
            errors = 1


        scanposition_progressbar['value'] = 100
        root.update_idletasks()

        scan_label.place_forget()
        scanposition_progressbar.place_forget()
        process_label.place_forget()
        wait = 0
        #global errors
        samplenum_label = Label(scanposition2_labelframe, bg='white', font=("Courier",10,'bold'))
        if(errors==1):
            errors=0
            err = messagebox.showerror('','L???i: ERR 01',icon = "error")
            samplenum_label['font']= ("Courier",12,'bold')
            samplenum_label['text'] = 'L???i: ERR 01'
            samplenum_label['fg'] = "red"
            samplenum_label.place(x=339,y=8)
        elif(errors==2 or errors==200):
            err = messagebox.showerror('','L???i: ERR 02', icon = "error")
            samplenum_label['font']= ("Courier",12,'bold')
            samplenum_label['text'] = 'L???i: ERR 02'
            samplenum_label['fg'] = "red"
            samplenum_label.place(x=339,y=8)
            err_labelframe = LabelFrame(scanposition1_labelframe, bg='black', font=("Courier",10,'bold'),width=306,height=356)

            def detail_click():
                if(detail_button['fg']=='blue'):
                    detail_button['fg']='black'
                    err_labelframe.place(x=246,y=3)
                    scanresult_labelframe.place_forget()
                    err_img = Image.open(path4 + '/mau.jpg')
                    err_crop = err_img.crop((x1-10, y1-10, x2+10, y2+10))
                    crop_width, crop_height = err_crop.size
                    scale_percent = 100
                    width = int(crop_width * scale_percent / 100)
                    height = int(crop_height * scale_percent / 100)
                    display_img = err_crop.resize((width,height))
                    err_display = ImageTk.PhotoImage(display_img)
                    err_label = Label(err_labelframe, image=err_display)
                    err_label.image = err_display
                    err_label.place(x=31,y=21)
                else:
                    detail_button['fg']='blue'
                    err_labelframe.place_forget()
                    scanresult_labelframe.place(x=244,y=4)

            def thread():
                th1 = Thread(target = next_click)
                th1.start()
            def next_click():
                global createclicked
                createclicked = 0
                scanposition_labelframe.place_forget()
                analysis()

            detail_button = Button(scanposition2_labelframe, font=("Courier",11),fg='blue', bg="white", activebackground = 'white', highlightbackground = 'white',text=">> Xem ???nh", height=1, width=8, borderwidth=0, command=detail_click)
            detail_button.place(x=347,y=30)
            if(errors==2):
                next_button = Button(scanposition2_labelframe, font=("Courier",10,'bold'), bg="lavender", text="Ti???p theo", height=3, width=11, borderwidth=0,command=thread)
                next_button.place(x=677,y=1)
            errors=0
        else:
            info = messagebox.showinfo('Ho??n th??nh','C?? th??? cho m???u v??o v?? ti???n h??nh ph??n t??ch.')
            samplenum_label['font']= ("Courier",12,'bold')
            samplenum_label['text'] = '???? KI???M TRA XONG !'
            samplenum_label['fg'] = "green4"
            samplenum_label.place(x=307,y=8)

            global thr_set
            thr_set = round(average_value*a+b,1)

            def thread():
                th1 = Thread(target = next_click)
                th1.start()

            def next_click():
                global createclicked
                createclicked = 0
                scanposition_labelframe.place_forget()
                analysis()

            next_button = Button(scanposition2_labelframe, font=("Courier",10,'bold'), bg="lavender", text="Ti???p theo", height=3, width=11, borderwidth=0,command=thread)
            next_button.place(x=677,y=1)

########################################################## SAMPLES POSITION - END ##################################################################

######################################################### SAMPLES ANALYSIS - START #################################################################
def analysis():
    global path0
    global path1
    global path2
    global path3
    global path4
    global path5

    global analysis_mode
    analysis_mode = 1

    global ser
    ser.flushInput()
    ser.flushOutput()

    global analysis_labelframe
    analysis_labelframe = LabelFrame(root, bg='white', width=800, height=480)
    analysis_labelframe.place(x=0,y=0)
    analysis1_labelframe = LabelFrame(analysis_labelframe, bg=atk.DEFAULT_COLOR, width=797, height=368)
    analysis1_labelframe.place(x=0,y=41)
    analysis2_labelframe = LabelFrame(analysis_labelframe, bg='white', width=797, height=67)
    analysis2_labelframe.place(x=0,y=409)
    title_labelframe = LabelFrame(analysis_labelframe, bg='dodger blue', width=796, height=40)
    title_labelframe.place(x=0,y=0)
    analysis_label = Label(analysis_labelframe, bg='dodger blue', text='PH??N T??CH M???U', font=("Courier",12,'bold'), width=20, height=1 )
    analysis_label.place(x=295,y=10)

    a1_labelframe = LabelFrame(analysis_labelframe, text="???nh ch???p", fg='lawn green', bg=atk.DEFAULT_COLOR, width=245, height=330)
    a1_labelframe.place(x=100,y=57)
    a2_labelframe = LabelFrame(analysis_labelframe, text="???nh ph??n t??ch", fg='lawn green', bg=atk.DEFAULT_COLOR, width=245, height=330)
    a2_labelframe.place(x=448,y=57)

    t_progressbar = atk.RadialProgressbar(a1_labelframe, fg='cyan')
    t_progressbar.place(x=69,y=112)
    t_progressbar.start()
    tprocess_label = Label(a1_labelframe, bg=atk.DEFAULT_COLOR, fg='white smoke', text='??ang x??? l?? !', font=("Courier",9,'bold'))
    tprocess_label.place(x=77,y=156)

    t_progressbar = atk.RadialProgressbar(a2_labelframe, fg='cyan')
    t_progressbar.place(x=69,y=112)
    t_progressbar.start()
    tprocess_label = Label(a2_labelframe, bg=atk.DEFAULT_COLOR, fg='white smoke', text='??ang x??? l?? !', font=("Courier",9,'bold'))
    tprocess_label.place(x=77,y=156)

    send_data = "P"
    ser.write(send_data.encode())
    print("Data send: ", send_data)

    process_label = Label(analysis_labelframe, text='??ang ph??n t??ch...', bg='white', font=("Courier",11))
    process_label.place(x=320,y=448)
    root.update_idletasks()
    scanposition_progressbar = ttk.Progressbar(root, orient = HORIZONTAL, style="green.Horizontal.TProgressbar", length = 200, mode = 'determinate')
    scanposition_progressbar.place(x=299,y=422)
    scanposition_progressbar['value'] = 5
    root.update_idletasks()

    global wait
    if(ser.in_waiting>0):
        receive_data = ser.readline().decode('utf-8').rstrip()
        print("Data received:", receive_data)
        scanposition_progressbar['value'] = 20
        root.update_idletasks()
        if(receive_data=='C'):
            wait = 1
    while(wait!=1):
        if(ser.in_waiting>0):
            receive_data = ser.readline().decode('utf-8').rstrip()
            print("Data received:", receive_data)
            scanposition_progressbar['value'] = 30
            root.update_idletasks()
            if(receive_data=='C'):
                wait = 1
                break

    global id_list
    while(wait==1):
        try:
            camera_capture(path1 + "/anh-chup.jpg")
        except Exception as e :
            error = messagebox.askquestion("L???i: "+ str(e), "B???n c?? mu???n tho??t ch????ng tr??nh ?", icon = "error")
            if(error=='yes'):
                root.destroy()

        global result_list
        result_list, result_img = process_image(path1 + "/anh-chup.jpg")

        scanposition_progressbar['value'] = 50
        root.update_idletasks()

        output = path2 + "/anh-xu-ly.jpg"
        cv2.imwrite(output, result_img)

        workbook = Workbook()
        sheet = workbook.active
        sheet["I11"] = thr_set
        sheet["A2"] = "A"
        sheet["A3"] = "B"
        sheet["A4"] = "C"
        sheet["A5"] = "D"
        sheet["A6"] = "E"
        sheet["A7"] = "F"
        sheet["A8"] = "G"
        sheet["A9"] = "H"
        sheet["B1"] = "1"
        sheet["C1"] = "2"
        sheet["D1"] = "3"
        sheet["E1"] = "4"
        sheet["F1"] = "5"
        sheet["G1"] = "6"
        for i in range(0,48):
            if(i<6):
                pos = str(chr(65+i+1)) + "2"
            if(i>=6 and i<12):
                pos = str(chr(65+i-5)) + "3"
            if(i>=12 and i<18):
                pos = str(chr(65+i-11)) + "4"
            if(i>=18 and i<24):
                pos = str(chr(65+i-17)) + "5"
            if(i>=24 and i<30):
                pos = str(chr(65+i-23)) + "6"
            if(i>=30 and i<36):
                pos = str(chr(65+i-29)) + "7"
            if(i>=36 and i<42):
                pos = str(chr(65+i-35)) + "8"
            if(i>=42):
                pos = str(chr(65+i-41)) + "9"

            if(id_list[i]=='N/A'):
                sheet[pos] = 'N/A'
            else:
                sheet[pos] = result_list[i]

        workbook.save(path3+"/gia-tri.xlsx")

        scanposition_progressbar['value'] = 65
        root.update_idletasks()

        if(server_on == 1):
            workbook1 = load_workbook("/home/pi/Desktop/Spotcheck ID/" + excel_file, keep_vba = True)
            sheet = workbook1.active
        else:
            workbook1 = load_workbook("/home/pi/Spotcheck/template.xlsm", keep_vba = True)
            sheet = workbook1.active

        sheet.protection.sheet = True
        sheet.protection.enable()

        scanposition_progressbar['value'] = 75
        root.update_idletasks()

        if(server_on == 0):
            sheet["C10"].protection = Protection(locked=False, hidden=False)
        #sheet["B7"].protection = Protection(locked=False, hidden=False)
        sheet["B8"].protection = Protection(locked=False, hidden=False)
        #sheet["B9"].protection = Protection(locked=False, hidden=False)

        font0 = Font(bold=False)
        font1 = Font(size='14', bold=True, color='00FF0000')
        font2 = Font(bold=True)
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        for i in range(12,60):
            sheet["B"+str(i)].font = font0
            sheet["D"+str(i)].font = font0

        img = Img('/home/pi/Spotcheck/logo.png')
        img.height = 39
        img.width = 215
        img.anchor = 'B2'
        sheet.add_image(img)

        sheet.merge_cells(start_row=5, start_column=2, end_row=5, end_column=6)
        sheet["B5"] = 'K???T QU??? CH???N ??O??N COVID-19'
        sheet["B5"].font = font1
        sheet.cell(row=5,column=2).alignment = Alignment(horizontal='center',vertical='center',wrapText=True)
        #global foldername
        sheet["B7"] = 'T??n t???p x??t nghi???m: ' + importfilename
        sheet["B7"].font = font2
        sheet['B8'] = 'Ng?????i th???c hi???n: '
        sheet["B8"].font = font2
        #global covid19dir_old
        sheet['B9'] = 'Ng??y th???c hi???n: ' + covid19dir_old[8:25]
        sheet["B9"].font = font2
        sheet['B60'] = 'Ghi ch??:'
        sheet["B60"].font = font2
        sheet['B61'] = '+ N/A: Tr???ng'
        sheet['B62'] = '+ OK: ?????t'
        sheet['C61'] = '+ L: Th???p'
        sheet['C62'] = '+ H: Cao'

        sheet.merge_cells(start_row=64, start_column=4, end_row=64, end_column=6)
        sheet.merge_cells(start_row=65, start_column=4, end_row=65, end_column=6)
        sheet['B64'] = '???K??? thu???t vi??n'
        sheet['B65'] = 'K?? t??n'
        sheet['D64'] = '???Tr?????ng ph??ng x??t nghi???m'
        sheet['D65'] = 'K?? t??n'
        sheet["B64"].font = font2
        sheet["D64"].font = font2
        sheet["B64"].protection = Protection(locked=False, hidden=False)
        sheet["D64"].protection = Protection(locked=False, hidden=False)
        sheet.cell(row=64,column=2).alignment = Alignment(horizontal='center',vertical='center',wrapText=True)
        sheet.cell(row=65,column=2).alignment = Alignment(horizontal='center',vertical='center',wrapText=True)
        sheet.cell(row=64,column=4).alignment = Alignment(horizontal='center',vertical='center',wrapText=True)
        sheet.cell(row=65,column=4).alignment = Alignment(horizontal='center',vertical='center',wrapText=True)

        for r in range(11,60):
            for c in range(2,7):
                sheet.cell(row=r,column=c).alignment = Alignment(horizontal='center',vertical='center',wrapText=True)
                sheet.cell(row=r,column=c).border = thin_border

        sheet.column_dimensions['B'].width = 26
        sheet.column_dimensions['C'].width = 12
        sheet.column_dimensions['D'].width = 12
        sheet.column_dimensions['E'].width = 12
        sheet.column_dimensions['F'].width = 12

        sheet.row_dimensions[11].height = 40

        sheet['B11'] = 'ID KH??CH H??NG'
        sheet["B11"].font = font2
        sheet["B11"].fill = PatternFill(start_color='00EFEFEF', end_color='00EFEFEF', fill_type='solid')
        sheet['C11'] = 'V??? tr?? m???u'
        sheet["C11"].font = font2
        sheet["C11"].fill = PatternFill(start_color='00EFEFEF', end_color='00EFEFEF', fill_type='solid')
        sheet['D11'] = 'K???t qu??? Spotcheck'
        sheet["D11"].font = font2
        sheet["D11"].fill = PatternFill(start_color='00EFEFEF', end_color='00EFEFEF', fill_type='solid')
        sheet['E11'] = 'K???t qu??? Gel'
        sheet["E11"].font = font2
        sheet["E11"].fill = PatternFill(start_color='00EFEFEF', end_color='00EFEFEF', fill_type='solid')
        sheet['F11'] = 'K???t lu???n'
        sheet["F11"].font = font2
        sheet["F11"].fill = PatternFill(start_color='00EFEFEF', end_color='00EFEFEF', fill_type='solid')

        for i in range (12,60):
            if(i<20):
                sheet['C'+str(i)] = str(chr(65+i-12)) + '1'
            if(i>=20 and i<28):
                sheet['C'+str(i)] = str(chr(65+i-20)) + '2'
            if(i>=28 and i<36):
                sheet['C'+str(i)] = str(chr(65+i-28)) + '3'
            if(i>=36 and i<44):
                sheet['C'+str(i)] = str(chr(65+i-36)) + '4'
            if(i>=44 and i<52):
                sheet['C'+str(i)] = str(chr(65+i-44)) + '5'
            if(i>=52):
                sheet['C'+str(i)] = str(chr(65+i-52)) + '6'

        c1=-6
        c2=-5
        c3=-4
        c4=-3
        c5=-2
        c6=-1
        for i in range(0,8):
            c1=c1+6
            sheet['B'+str(i+12)] = id_list[c1]
            if(id_list[c1]=='N/A'):
                sheet['D'+str(i+12)] = 'N/A'
            else:
                if(result_list[c1] > 0.8*float(thr_set)*1.05):
                    sheet['D'+str(i+12)] = 'H'
                    sheet['D'+str(i+12)].fill = PatternFill(start_color='00FFCC00', end_color='00FFCC00', fill_type='solid')
                elif(result_list[c1] < 0.8*float(thr_set)*0.95):
                    sheet['D'+str(i+12)] = 'L'
                    sheet['D'+str(i+12)].fill = PatternFill(start_color='00FFFF00', end_color='00FFFF00', fill_type='solid')
                else:
                    sheet['D'+str(i+12)] = 'OK'
                    sheet['D'+str(i+12)].fill = PatternFill(start_color='0000FF00', end_color='0000FF00', fill_type='solid')


            sheet['E'+str(i+12)].protection = Protection(locked=False, hidden=False)
            sheet['F'+str(i+12)].protection = Protection(locked=False, hidden=False)

            c2=c2+6
            sheet['B'+str(i+20)] = id_list[c2]
            if(id_list[c2]=='N/A'):
                sheet['D'+str(i+20)] = 'N/A'
            else:
                if(result_list[c2] > 0.8*float(thr_set)*1.05):
                    sheet['D'+str(i+20)] = 'H'
                    sheet['D'+str(i+20)].fill = PatternFill(start_color='00FFCC00', end_color='00FFCC00', fill_type='solid')
                elif(result_list[c2] < 0.8*float(thr_set)*0.95):
                    sheet['D'+str(i+20)] = 'L'
                    sheet['D'+str(i+20)].fill = PatternFill(start_color='00FFFF00', end_color='00FFFF00', fill_type='solid')
                else:
                    sheet['D'+str(i+20)] = 'OK'
                    sheet['D'+str(i+20)].fill = PatternFill(start_color='0000FF00', end_color='0000FF00', fill_type='solid')

            sheet['E'+str(i+20)].protection = Protection(locked=False, hidden=False)
            sheet['F'+str(i+20)].protection = Protection(locked=False, hidden=False)

            c3=c3+6
            sheet['B'+str(i+28)] = id_list[c3]
            if(id_list[c3]=='N/A'):
                sheet['D'+str(i+28)] = 'N/A'
            else:
                if(result_list[c3] > 0.8*float(thr_set)*1.05):
                    sheet['D'+str(i+28)] = 'H'
                    sheet['D'+str(i+28)].fill = PatternFill(start_color='00FFCC00', end_color='00FFCC00', fill_type='solid')
                elif(result_list[c3] < 0.8*float(thr_set)*0.95):
                    sheet['D'+str(i+28)] = 'L'
                    sheet['D'+str(i+28)].fill = PatternFill(start_color='00FFFF00', end_color='00FFFF00', fill_type='solid')
                else:
                    sheet['D'+str(i+28)] = 'OK'
                    sheet['D'+str(i+28)].fill = PatternFill(start_color='0000FF00', end_color='0000FF00', fill_type='solid')

            sheet['E'+str(i+28)].protection = Protection(locked=False, hidden=False)
            sheet['F'+str(i+28)].protection = Protection(locked=False, hidden=False)

            c4=c4+6
            sheet['B'+str(i+36)] = id_list[c4]
            if(id_list[c4]=='N/A'):
                sheet['D'+str(i+36)] = 'N/A'
            else:
                if(result_list[c4] > 0.8*float(thr_set)*1.05):
                    sheet['D'+str(i+36)] = 'H'
                    sheet['D'+str(i+36)].fill = PatternFill(start_color='00FFCC00', end_color='00FFCC00', fill_type='solid')
                elif(result_list[c4] < 0.8*float(thr_set)*0.95):
                    sheet['D'+str(i+36)] = 'L'
                    sheet['D'+str(i+36)].fill = PatternFill(start_color='00FFFF00', end_color='00FFFF00', fill_type='solid')
                else:
                    sheet['D'+str(i+36)] = 'OK'
                    sheet['D'+str(i+36)].fill = PatternFill(start_color='0000FF00', end_color='0000FF00', fill_type='solid')

            sheet['E'+str(i+36)].protection = Protection(locked=False, hidden=False)
            sheet['F'+str(i+36)].protection = Protection(locked=False, hidden=False)

            c5=c5+6
            sheet['B'+str(i+44)] = id_list[c5]
            if(id_list[c5]=='N/A'):
                sheet['D'+str(i+44)] = 'N/A'
            else:
                if(result_list[c4] > 0.8*float(thr_set)*1.05):
                    sheet['D'+str(i+44)] = 'H'
                    sheet['D'+str(i+44)].fill = PatternFill(start_color='00FFCC00', end_color='00FFCC00', fill_type='solid')
                elif(result_list[c4] < 0.8*float(thr_set)*0.95):
                    sheet['D'+str(i+44)] = 'L'
                    sheet['D'+str(i+44)].fill = PatternFill(start_color='00FFFF00', end_color='00FFFF00', fill_type='solid')
                else:
                    sheet['D'+str(i+44)] = 'OK'
                    sheet['D'+str(i+44)].fill = PatternFill(start_color='0000FF00', end_color='0000FF00', fill_type='solid')

            sheet['E'+str(i+44)].protection = Protection(locked=False, hidden=False)
            sheet['F'+str(i+44)].protection = Protection(locked=False, hidden=False)

            c6=c6+6
            sheet['B'+str(i+52)] = id_list[c6]
            if(id_list[c6]=='N/A'):
                sheet['D'+str(i+52)] = 'N/A'
            else:
                if(result_list[c6] > 0.8*float(thr_set)*1.05):
                    sheet['D'+str(i+52)] = 'H'
                    sheet['D'+str(i+52)].fill = PatternFill(start_color='00FFCC00', end_color='00FFCC00', fill_type='solid')
                elif(result_list[c6] < 0.8*float(thr_set)*0.95):
                    sheet['D'+str(i+52)] = 'L'
                    sheet['D'+str(i+52)].fill = PatternFill(start_color='00FFFF00', end_color='00FFFF00', fill_type='solid')
                else:
                    sheet['D'+str(i+52)] = 'OK'
                    sheet['D'+str(i+52)].fill = PatternFill(start_color='0000FF00', end_color='0000FF00', fill_type='solid')

            sheet['E'+str(i+52)].protection = Protection(locked=False, hidden=False)
            sheet['F'+str(i+52)].protection = Protection(locked=False, hidden=False)

        sheet.print_area = 'A1:G70'
        workbook1.save("/home/pi/Desktop/Ket Qua Phan Tich/Probe/" + importfilename + ".xlsm")

        scanposition_progressbar['value'] = 80
        root.update_idletasks()

        if(os.path.exists("/home/pi/Desktop/Spotcheck ID/" + excel_file)):
            try:
                shutil.move("/home/pi/Desktop/Spotcheck ID/" + excel_file,"/home/pi/Desktop/Spotcheck ID/Spotcheck ID - Old")
            except:
                pass
        else:
            pass

        a1 = Image.open(path1 + '/anh-chup.jpg')
        a1_crop = a1.crop((x1-10, y1-10, x2+10, y2+10))
        crop_width, crop_height = a1_crop.size
        scale_percent = 100
        width = int(crop_width * scale_percent / 100)
        height = int(crop_height * scale_percent / 100)
        display_img = a1_crop.resize((width,height))
        a1_display = ImageTk.PhotoImage(display_img)
        a1_label = Label(a1_labelframe, image=a1_display)
        a1_label.image = a1_display
        a1_label.place(x=3,y=1)

        a2 = Image.open(path2 + '/anh-xu-ly.jpg')
        a2_crop = a2.crop((x1-10, y1-10, x2+10, y2+10))
        crop_width, crop_height = a2_crop.size
        scale_percent = 100
        width = int(crop_width * scale_percent / 100)
        height = int(crop_height * scale_percent / 100)
        display_img = a2_crop.resize((width,height))
        a2_display = ImageTk.PhotoImage(display_img)
        a2_label = Label(a2_labelframe, image=a2_display)
        a2_label.image = a2_display
        a2_label.place(x=3,y=1)

        if(server_on==1):
            try:
                ftp = FTP(ftp_ip, ftp_user, ftp_password, timeout=30)
                ftp.cwd(ftp_folder + 'Processed_Data')
                file = open("/home/pi/Desktop/Ket Qua Phan Tich/Probe/" + importfilename + ".xlsm",'rb')
                ftp.storbinary('STOR ' + importfilename + ".xlsm", file)
                ftp.quit()
            except Exception as e :
                error = messagebox.showwarning("C?? l???i x???y ra khi ?????ng b??? server !",str(e))
                if(error=='ok'):
                    pass

        scanposition_progressbar['value'] = 90
        root.update_idletasks()
        sleep(1)
        scanposition_progressbar['value'] = 100
        root.update_idletasks()

        process_label.place_forget()
        scanposition_progressbar.place_forget()

        def thr():
            th2 = Thread(target = viewresult_click)
            th2.start()
        def viewresult_click():
            global analysis_mode
            analysis_mode = 0
            a1_labelframe.place_forget()
            a2_labelframe.place_forget()
            analysis1_labelframe.place_forget()
            analysis2_labelframe.place_forget()
            viewresult_button.place_forget()
            analysis_label['text']="K???T QU??? PH??N T??CH"

            annotate_labelframe = LabelFrame(analysis_labelframe, bg='white', width=380, height=305)
            annotate_labelframe.place(x=360,y=66)
            root.update_idletasks()

            negative_label = Label(annotate_labelframe, bg='green3', width=4, height=2)
            negative_label.place(x=60,y=32)
            negativetext_label = Label(annotate_labelframe, bg='white', text='    (OK)                ?????T', height=2)
            negativetext_label.place(x=130,y=32)
            positive_label = Label(annotate_labelframe, bg='yellow', width=4, height=2)
            positive_label.place(x=60,y=82)
            positivetext_label = Label(annotate_labelframe, bg='white', text='    (L)                   TH???P', height=2)
            positivetext_label.place(x=130,y=82)
            positive_label = Label(annotate_labelframe, bg='orange', width=4, height=2)
            positive_label.place(x=60,y=132)
            positivetext_label = Label(annotate_labelframe, bg='white', text='    (H)                  CAO', height=2)
            positivetext_label.place(x=130,y=132)
            na_label = Label(annotate_labelframe, bg='white smoke', width=4, height=2)
            na_label.place(x=60,y=232)
            natext_label = Label(annotate_labelframe, bg='white', text='    (N/A)               TR???NG', height=2)
            natext_label.place(x=130,y=232)
            root.update_idletasks()

            result_labelframe = LabelFrame(analysis_labelframe, bg='ghost white', width=600,height = 307)
            result_labelframe.place(x=104,y=110)
            row_labelframe = LabelFrame(analysis_labelframe, bg='ghost white', width=600,height = 50)
            row_labelframe.place(x=104,y=66)
            column_labelframe = LabelFrame(analysis_labelframe, bg='ghost white', width=50,height = 307)
            column_labelframe.place(x=62,y=110)
            root.update_idletasks()

            row_label = [0,0,0,0,0,0]
            for i in range (0,6):
                row_text = i+1
                row_label[i] = Label(row_labelframe, text=row_text, bg='grey94', width=4, height=2)
                row_label[i].grid(row=0,column=i,padx=2,pady=2)

            column_label = [0,0,0,0,0,0,0,0]
            for i in range (0,8):
                if(i==0):
                    column_text = 'A'
                if(i==1):
                    column_text = 'B'
                if(i==2):
                    column_text = 'C'
                if(i==3):
                    column_text = 'D'
                if(i==4):
                    column_text = 'E'
                if(i==5):
                    column_text = 'F'
                if(i==6):
                    column_text = 'G'
                if(i==7):
                    column_text = 'H'
                column_label[i] = Label(column_labelframe, text=column_text, bg='grey94', width=4, height=2)
                column_label[i].grid(row=i,column=0,padx=2,pady=2)

            label = list(range(48))
            def result_table(range_a,range_b, row_value):
                j=-1
                global pos_result
                for i in range(range_a, range_b):
                    j+=1
                    if(id_list[i]=='N/A'):
                        label[i] = Label(result_labelframe, bg='white smoke', text='N/A', width=4, height=2)
                        label[i].grid(row=row_value,column=j,padx=2,pady=2)
                    else:
                        if(result_list[i] > 0.8*float(thr_set)*1.05):
                            label[i] = Label(result_labelframe, bg='orange', text='H', width=4, height=2)
                            label[i].grid(row=row_value,column=j,padx=2,pady=2)
                        elif(result_list[i] < 0.8*float(thr_set)*0.95):
                            label[i] = Label(result_labelframe, bg='yellow', text='L', width=4, height=2)
                            label[i].grid(row=row_value,column=j,padx=2,pady=2)
                        else:
                            label[i] = Label(result_labelframe, bg='green3', text='OK', width=4, height=2)
                            label[i].grid(row=row_value,column=j,padx=2,pady=2)

            result_table(0,6,0)
            result_table(6,12,1)
            result_table(12,18,2)
            result_table(18,24,3)
            result_table(24,30,4)
            result_table(30,36,5)
            result_table(36,42,6)
            result_table(42,48,7)

            root.update_idletasks()

            def finish_click():
                msgbox = messagebox.askquestion('Ket thuc chuong trinh','B???n c?? mu???n quay l???i ?', icon = 'question')
                if(msgbox=='yes'):
                    # for i in range (0,48):
                    #     label[i]['text'] = str('%.1f'%t3_result[i])
                    # root.update_idletasks()
                    # subprocess.call(["scrot",path3+"/gia-tri.jpg"])
                    # sleep(1)
                    global foldername
                    global covid19clicked
                    foldername = ""
                    covid19clicked = 0
                    analysis_labelframe.place_forget()
                    global wait, analysis_mode
                    wait=0
                    analysis_mode = 0
                    mainscreen()

            finish_button = Button(analysis_labelframe, bg="lawn green", text="Ho??n th??nh", height=3, width=15, borderwidth=0, command=finish_click)
            finish_button.place(x=477,y=386)

            root.update_idletasks()

            subprocess.call(["scrot",path0+"/ket-qua.jpg"])

        viewresult_button = Button(analysis2_labelframe, bg="dodger blue", text="K???t qu???", height=3, width=12, borderwidth=0, command=thr)
        viewresult_button.place(x=335,y=0)
        wait = 0
########################################################## SAMPLES ANALYSIS - END ##################################################################

############################################################## WARNING - START #####################################################################
# def warning(channel):
#     global warning_value
#     if(warning_value==1):
#         warning_label = Label(mainscreen_labelframe, bg='white', fg='white', text='H??? th???ng ??ang t???n nhi???t, kh??ng ?????t m???u v??o l??c n??y !', font=("Courier", 13, 'bold'))
#         warning_label.place(x=220,y=450)
#         warning_value = 0
#         print("Warning:", warning_value)
#     else:
#         warning_label = Label(mainscreen_labelframe, bg='red',text='H??? th???ng ??ang t???n nhi???t, kh??ng ?????t m???u v??o l??c n??y !', font=("Courier", 13, 'bold'))
#         warning_label.place(x=220,y=450)
#         warning_value = 1
#         print("Warning:", warning_value)

# GPIO.setmode(GPIO.BCM)
# GPIO.setup(16, GPIO.IN, pull_up_down=GPIO.PUD_DOWN)
# GPIO.add_event_detect(16,GPIO.FALLING,callback=warning)
############################################################### WARNING - END ######################################################################

############################################################### LOOP - START #######################################################################
if(start_trial==1):
    trial()
else:
    mainscreen()

root.mainloop()
################################################################ LOOP - END ########################################################################