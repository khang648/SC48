#!/usr/bin/python3
########################################################### IMPORT MODULES - START #################################################################
from tkinter import *
from tkinter import messagebox
import time
from time import sleep, gmtime, strftime
from picamera import PiCamera
import cv2
import numpy as np
from tkinter import filedialog
from PIL import ImageTk, Image
import serial
from functools import partial
import math
from fractions import Fraction
from threading import *
import os
from tkinter import ttk
import awesometkinter as atk
import tkinter.font as font
import openpyxl
import subprocess
import shutil
import RPi.GPIO as GPIO
from ftplib import FTP
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Protection
from openpyxl.styles.borders import Border, Side
from openpyxl.drawing.image import Image as Img
from datetime import *
############################################################ IMPORT MODULES - END ##################################################################

########################################################## GLOBAL VARIABLE - START #################################################################
wait = 0
result_list = list(range(48))
covid19dir_old = ""
start_point = (0,0)
end_point = (0,0)
path = ''
path0=''

hs = list(range(48))
workbook = openpyxl.load_workbook('/home/pi/Spotcheck/coefficient.xlsx')
sheet = workbook.active
for i in range(0,48):
    if(i<6):
        pos = str(chr(65+i+1)) + "2"
    if(i>=6 and i<12):
        pos = str(chr(65+i-5)) + "3"
    if(i>=12 and i<18):
        pos = str(chr(65+i-11)) + "4"
    if(i>=18 and i<24):
        pos = str(chr(65+i-17)) + "5"
    if(i>=24 and i<30):
        pos = str(chr(65+i-23)) + "6"
    if(i>=30 and i<36):
        pos = str(chr(65+i-29)) + "7"
    if(i>=36 and i<42):
        pos = str(chr(65+i-35)) + "8"
    if(i>=42):
        pos = str(chr(65+i-41)) + "9"
    hs[i] = float(sheet[pos].value)

fr1 = open("/home/pi/Spotcheck/coordinates1.txt","r")
x1 = int(fr1.readline())
y1 = int(fr1.readline())
x2 = int(fr1.readline())
y2 = int(fr1.readline())
########################################################### GLOBAL VARIABLE - END ##################################################################

########################################################## MAIN WINDOW INIT - START ################################################################
root = Tk()
root.title(" ")
root.geometry('1024x600')
root.configure(background = "white")
root.attributes('-fullscreen', True)
root.resizable(False,False)
def disable_event():
    pass
root.protocol("WM_DELETE_WINDOW", disable_event)
s = ttk.Style()
s.theme_use('clam')
########################################################### MAIN WINDOW INIT - END #################################################################

############################################################ CAMERA INIT - START ###################################################################
def camera_capture(output):
    global stop_click
    stop_click = 0
    camera = PiCamera(framerate=Fraction(1,6), sensor_mode=3)
    camera.rotation = 180
    camera.iso = 200
    sleep(2)
    camera.shutter_speed = 6000000
    camera.exposure_mode = 'off'
    camera.capture(output)
    camera.close()
############################################################# CAMERA INIT - END ####################################################################

############################################################ SERIAL INIT - START ###################################################################
ser = serial.Serial(
    port = '/dev/serial0',
    baudrate = 115200,
    parity = serial.PARITY_NONE,
    stopbits = serial.STOPBITS_ONE,
    bytesize = serial.EIGHTBITS,
    timeout = 1
)
############################################################# SERIAL INIT - END ####################################################################

######################################################### SORTING CONTOURS - START #################################################################
def sorting_y(contour):
    rect_y = cv2.boundingRect(contour)
    return rect_y[1]
def sorting_x(contour):
    rect_x = cv2.boundingRect(contour)
    return rect_x[0]
def sorting_xy(contour):
    rect_xy = cv2.boundingRect(contour)
    return math.sqrt(math.pow(rect_xy[0],2) + math.pow(rect_xy[1],2))
########################################################## SORTING CONTOURS - END ##################################################################

########################################################## IMAGE ANALYSIS - START ##################################################################
def process_image(image_name, color, start_point=(x1,y1), end_point=(x2,y2)):
    image = cv2.imread(image_name)
    blur_img = cv2.GaussianBlur(image.copy(), (35,35), 0)
    gray_img = cv2.cvtColor(blur_img, cv2.COLOR_BGR2GRAY)
    thresh, binary_img = cv2.threshold(gray_img.copy(), 30, maxval=255, type=cv2.THRESH_BINARY)
    contours, hierarchy = cv2.findContours(binary_img, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    print("Number of contours: " + str(len(contours)))

    contours.sort(key=lambda data:sorting_xy(data))

    contour_img = np.zeros_like(gray_img)
    contour_img = cv2.rectangle(contour_img, start_point, end_point, (255,255,255), -1)
    rect_w = end_point[0] - start_point[0]
    rect_h = end_point[1] - start_point[1]
    cell_w = round(rect_w/6)
    cell_h = round(rect_h/8)
    for i in range(1,6):
        contour_img = cv2.line(contour_img, (start_point[0]+i*cell_w,start_point[1]), (start_point[0]+i*cell_w,end_point[1]),(0,0,0), 4)
    for i in range(1,8):
        contour_img = cv2.line(contour_img, (start_point[0],start_point[1]+i*cell_h), (end_point[0],start_point[1]+i*cell_h),(0,0,0), 4)

    thresh1 , binary1_img = cv2.threshold(contour_img, 250, maxval=255, type=cv2.THRESH_BINARY)
    contours1, hierarchy1 = cv2.findContours(binary1_img, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_NONE)

    contours1.sort(key=lambda data:sorting_y(data))
    contours1_h1 = contours1[0:6]
    contours1_h2 = contours1[6:12]
    contours1_h3 = contours1[12:18]
    contours1_h4 = contours1[18:24]
    contours1_h5 = contours1[24:30]
    contours1_h6 = contours1[30:36]
    contours1_h7 = contours1[36:42]
    contours1_h8 = contours1[42:48]
    contours1_h1.sort(key=lambda data:sorting_x(data))
    contours1_h2.sort(key=lambda data:sorting_x(data))
    contours1_h3.sort(key=lambda data:sorting_x(data))
    contours1_h4.sort(key=lambda data:sorting_x(data))
    contours1_h5.sort(key=lambda data:sorting_x(data))
    contours1_h6.sort(key=lambda data:sorting_x(data))
    contours1_h7.sort(key=lambda data:sorting_x(data))
    contours1_h8.sort(key=lambda data:sorting_x(data))

    sorted_contours1 = contours1_h1 + contours1_h2 + contours1_h3 + contours1_h4 + contours1_h5 + contours1_h6 + contours1_h7 + contours1_h8

    list_intensities = []
    sum_intensities = []
    result_list = list(range(48))
    area = list(range(48))

#BGR
    blur1_img = cv2.GaussianBlur(image.copy(), (25,25), 0)
    tmp_list = list(range(48))
    list_bgrvalue = []
    list_index = list(range(48))
    for i in range(len(sorted_contours1)):
        list_index[i] = []
        cimg = np.zeros_like(gray_img)
        cv2.drawContours(cimg, sorted_contours1, i, color = 255, thickness = -1)
        pts = np.where(cimg == 255)
        list_bgrvalue.append(blur1_img[pts[0], pts[1]])
        for j in range(len(list_bgrvalue[i])):
             list_index[i].append(round((list_bgrvalue[i][j][color])))
        list_index[i].sort()
        list_intensities.append(sum(list_index[i][len(list_index[i])-250:]))
        area[i]= cv2.contourArea(sorted_contours1[i])
        tmp_list[i] = list_intensities[i]/1000
        result_list[i] = round(tmp_list[i],1)

#Nhân hệ số
    global hs
    for i in range(1,7):
        result_list[6*i+1]=round(result_list[6*i+1]*(1-(0.02*round(result_list[6*i]/70) + 0.02*round(result_list[6*i+2]/70) + 0.02*round(result_list[6*i-5]/70) + 0.02*round(result_list[6*i+7]/70) + 0.003*round(result_list[6*i-6]/70) + 0.003*round(result_list[6*i-4]/70) + 0.003*round(result_list[6*i+6]/70) + 0.003*round(result_list[6*i+8]/70)+ 0.006*round(result_list[6*i+3]/70))),1)
        result_list[6*i+2]=round(result_list[6*i+2]*(1-(0.02*round(result_list[6*i+1]/70) + 0.02*round(result_list[6*i+3]/70) + 0.02*round(result_list[6*i-4]/70) + 0.02*round(result_list[6*i+8]/70) + 0.003*round(result_list[6*i-5]/70) + 0.003*round(result_list[6*i-3]/70) + 0.003*round(result_list[6*i+7]/70) + 0.003*round(result_list[6*i+9]/70)+ 0.006*round(result_list[6*i+4]/70)+ 0.006*round(result_list[6*i]/70))),1)
        result_list[6*i+3]=round(result_list[6*i+3]*(1-(0.02*round(result_list[6*i+2]/70) + 0.02*round(result_list[6*i+4]/70) + 0.02*round(result_list[6*i-3]/70) + 0.02*round(result_list[6*i+9]/70) + 0.003*round(result_list[6*i-4]/70) + 0.003*round(result_list[6*i-2]/70) + 0.003*round(result_list[6*i+8]/70) + 0.003*round(result_list[6*i+10]/70)+ 0.006*round(result_list[6*i+5]/70)+ 0.006*round(result_list[6*i+1]/70))),1)
        result_list[6*i+4]=round(result_list[6*i+4]*(1-(0.02*round(result_list[6*i+3]/70) + 0.02*round(result_list[6*i+5]/70) + 0.02*round(result_list[6*i-2]/70) + 0.02*round(result_list[6*i+10]/70) + 0.003*round(result_list[6*i-3]/70) + 0.003*round(result_list[6*i-1]/70) + 0.003*round(result_list[6*i+9]/70) + 0.003*round(result_list[6*i+11]/70)+ 0.006*round(result_list[6*i+2]/70))),1)

        result_list[6*i]=round(result_list[6*i]*(1-(0.02*round(result_list[6*i+1]/70) + 0.015*round(result_list[6*i-6]/70) + 0.015*round(result_list[6*i+6]/70) + 0.003*round(result_list[6*i-5]/70) + 0.003*round(result_list[6*i+7]/70)+0.006*round(result_list[6*i+2]/70))),1)
        result_list[6*i+5]=round(result_list[6*i+5]*(1-(0.02*round(result_list[6*i+4]/70) + 0.015*round(result_list[6*i-1]/70) + 0.015*round(result_list[6*i+11]/70) + 0.003*round(result_list[6*i-2]/70) + 0.003*round(result_list[6*i+10]/70)+0.006*round(result_list[6*i+3]/70))),1)

    for i in range(1,5):
        result_list[i]=round(result_list[i]*(1-(0.02*round(result_list[i-1]/70) + 0.02*round(result_list[i+1]/70) + 0.015*round(result_list[i+6]/70)+ 0.003*round(result_list[i+5]/70) + 0.003*round(result_list[i+7]/70))),1)
        result_list[i+42]=round(result_list[i+42]*(1-(0.02*round(result_list[i+41]/70) + 0.02*round(result_list[i+43]/70) + 0.015*round(result_list[i+36]/70)+ 0.003*round(result_list[i+35]/70) + 0.003*round(result_list[i+37]/70))),1)

    result_list[0]=round(result_list[0]*(1-(0.015*round(result_list[1]/70) + 0.015*round(result_list[6]/70))),1)
    result_list[5]=round(result_list[5]*(1-(0.015*round(result_list[4]/70) + 0.015*round(result_list[11]/70))),1)
    result_list[42]=round(result_list[42]*(1-(0.015*round(result_list[43]/70) + 0.015*round(result_list[36]/70))),1)
    result_list[47]=round(result_list[47]*(1-(0.015*round(result_list[46]/70) + 0.015*round(result_list[41]/70))),1)


    for i in range(len(sorted_contours1)):
        if(result_list[i]>99):
            result_list[i]=99

    for i in range(len(sorted_contours1)):
        if ((i!=0) and ((i+1)%6==0)):
            print('%.1f'%(result_list[i]))
        else:
            print('%.1f'%(result_list[i]), end = ' | ')

    blurori_img = cv2.GaussianBlur(image.copy(), (25,25), 0)
    for i in range(len(sorted_contours1)):
        cv2.drawContours(blurori_img, sorted_contours1, i, (0,255,0), thickness = 2)

    workbook = load_workbook(path0 + "bgr.xlsx", keep_vba = True)
    sheet = workbook.active
    for i in range(0,48):
        if(color==0):
            if(i<6):
                pos = str(chr(65+i+1)) + "2"
            if(i>=6 and i<12):
                pos = str(chr(65+i-5)) + "3"
            if(i>=12 and i<18):
                pos = str(chr(65+i-11)) + "4"
            if(i>=18 and i<24):
                pos = str(chr(65+i-17)) + "5"
            if(i>=24 and i<30):
                pos = str(chr(65+i-23)) + "6"
            if(i>=30 and i<36):
                pos = str(chr(65+i-29)) + "7"
            if(i>=36 and i<42):
                pos = str(chr(65+i-35)) + "8"
            if(i>=42):
                pos = str(chr(65+i-41)) + "9"

        if(color==1):
            if(i<6):
                pos = str(chr(65+i+1)) + "11"
            if(i>=6 and i<12):
                pos = str(chr(65+i-5)) + "12"
            if(i>=12 and i<18):
                pos = str(chr(65+i-11)) + "13"
            if(i>=18 and i<24):
                pos = str(chr(65+i-17)) + "14"
            if(i>=24 and i<30):
                pos = str(chr(65+i-23)) + "15"
            if(i>=30 and i<36):
                pos = str(chr(65+i-29)) + "16"
            if(i>=36 and i<42):
                pos = str(chr(65+i-35)) + "17"
            if(i>=42):
                pos = str(chr(65+i-41)) + "18"

        if(color==2):
            if(i<6):
                pos = str(chr(65+i+1)) + "20"
            if(i>=6 and i<12):
                pos = str(chr(65+i-5)) + "21"
            if(i>=12 and i<18):
                pos = str(chr(65+i-11)) + "22"
            if(i>=18 and i<24):
                pos = str(chr(65+i-17)) + "23"
            if(i>=24 and i<30):
                pos = str(chr(65+i-23)) + "24"
            if(i>=30 and i<36):
                pos = str(chr(65+i-29)) + "25"
            if(i>=36 and i<42):
                pos = str(chr(65+i-35)) + "26"
            if(i>=42):
                pos = str(chr(65+i-41)) + "27"

        sheet[pos] = result_list[i]
        workbook.save(path0 + "bgr.xlsx")

    return (result_list, blurori_img)
########################################################### IMAGE ANALYSIS - END ###################################################################

############################################################ MAIN SCREEN - START ###################################################################
def mainscreen():
    mainscreen_labelframe = LabelFrame(root, bg='white', width=800, height=600)
    mainscreen_labelframe.place(x=0,y=0)

    def capture_click():
        rsfile = filedialog.askopenfilename(initialdir='/home/pi/Spotcheck Ket Qua/Probe/',filetypes=[('jpg file','*.jpg')])
        if rsfile is not None:
            capture_button.place_forget()
            exit_button.place_forget()

            folder_name = strftime("BGR %m-%d-%Y %H.%M.%S")
            global path0
            path0= os.path.join("/home/pi/Desktop/Test/", folder_name + "/")
            os.mkdir(path0)

            workbook = Workbook()
            sheet = workbook.active
            sheet["A2"] = "BLUE"
            sheet["A11"] = "RED"
            sheet["A20"] = "GREEN"
            workbook.save(path0 + "bgr.xlsx")

            process_label = Label(mainscreen_labelframe, text='Đang xử lý...', bg='white', font=("Courier",13,'bold'))
            process_label.place(x=330,y=350)
            root.update_idletasks()
            scanposition_progressbar = ttk.Progressbar(mainscreen_labelframe, orient = HORIZONTAL, style="green.Horizontal.TProgressbar", length = 200, mode = 'determinate')
            scanposition_progressbar.place(x=290,y=310)
            scanposition_progressbar['value'] = 30
            root.update_idletasks()

            result_list, result_img = process_image(rsfile,0)
            result_list, result_img = process_image(rsfile,1)
            result_list, result_img = process_image(rsfile,2)



            scanposition_progressbar['value'] = 50
            root.update_idletasks()
            sleep(1)
            scanposition_progressbar['value'] = 70
            root.update_idletasks()
            sleep(1)
            scanposition_progressbar['value'] = 99
            root.update_idletasks()

            process_label.place_forget()
            scanposition_progressbar.place_forget()

            mainscreen()

    def exit_click():
        msg = messagebox.askquestion('','Thoát hả bro?', icon = 'question')
        if(msg=='yes'):
            root.destroy()


    capture_button = Button(mainscreen_labelframe, bg="dodger blue", activebackground="white", text="LOAD", font=("Courier",14,'bold'), borderwidth=0, height=3, width=15,command=capture_click)
    capture_button.place(x=450,y=300)
    exit_button = Button(mainscreen_labelframe, bg="red", activebackground="white", text="EXIT", font=("Courier",14,'bold'), borderwidth=0, height=3, width=15, command=exit_click)
    exit_button.place(x=150,y=300)

############################################################# MAIN SCREEN - END ####################################################################

############################################################### LOOP - START #######################################################################
while True:
    mainscreen()
    root.mainloop()
################################################################ LOOP - END ########################################################################